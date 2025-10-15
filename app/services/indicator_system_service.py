# app/services/indicator_system_service.py
"""
指标体系建设服务层
处理指标体系的业务逻辑
"""
from typing import List, Optional, Dict, Any, Tuple
from sqlalchemy import select, and_, or_, func, desc
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from loguru import logger
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from app.models.indicator_system import IndicatorSystem, IndicatorAssetRelation
from app.models.data_asset import DataAsset
from app.schemas.indicator_system import (
    IndicatorSystemCreate, IndicatorSystemUpdate,
    IndicatorAssetRelationCreate, IndicatorStatistics
)


class IndicatorSystemService:
    """指标体系服务类"""

    # Excel模板表头定义（中文名）
    EXCEL_HEADERS_CN = [
        "来源系统", "业务领域", "业务主题", "指标类别", "指标名称", "指标说明", "备注",
        "指标类型", "数据标准技术分类", "数据类型", "数据长度", "数据格式",
        "权责部门", "采集频率", "采集时点", "共享类型", "开放属性"
    ]

    # Excel模板表头定义（英文字段名）
    EXCEL_HEADERS_EN = [
        "source_system", "business_domain", "business_theme", "indicator_category",
        "indicator_name", "indicator_description", "remark",
        "indicator_type", "tech_classification", "data_type", "data_length", "data_format",
        "responsible_dept", "collection_frequency", "collection_time", "share_type", "open_attribute"
    ]

    async def create_indicator(
            self,
            db: AsyncSession,
            indicator_data: IndicatorSystemCreate,
            creator: Optional[str] = None
    ) -> IndicatorSystem:
        """
        创建指标

        Args:
            db: 数据库会话
            indicator_data: 指标创建数据
            creator: 创建人

        Returns:
            创建的指标对象
        """
        try:
            # 创建指标对象
            indicator = IndicatorSystem(
                source_system=indicator_data.source_system,
                business_domain=indicator_data.business_domain,
                business_theme=indicator_data.business_theme,
                indicator_category=indicator_data.indicator_category,
                indicator_name=indicator_data.indicator_name,
                indicator_description=indicator_data.indicator_description,
                remark=indicator_data.remark,
                indicator_type=indicator_data.indicator_type,
                tech_classification=indicator_data.tech_classification,
                data_type=indicator_data.data_type,
                data_length=indicator_data.data_length,
                data_format=indicator_data.data_format,
                responsible_dept=indicator_data.responsible_dept,
                collection_frequency=indicator_data.collection_frequency,
                collection_time=indicator_data.collection_time,
                share_type=indicator_data.share_type,
                open_attribute=indicator_data.open_attribute,
                tags=indicator_data.tags,
                is_active=True,
                created_by=creator,
                updated_by=creator
            )

            db.add(indicator)
            await db.commit()
            await db.refresh(indicator)

            logger.info(f"创建指标成功: {indicator.indicator_name} (ID: {indicator.id})")
            return indicator

        except Exception as e:
            await db.rollback()
            logger.error(f"创建指标失败: {e}")
            raise ValueError(f"创建指标失败: {str(e)}")

    async def get_indicator_by_id(
            self,
            db: AsyncSession,
            indicator_id: int,
            load_relations: bool = False
    ) -> Optional[IndicatorSystem]:
        """
        根据ID获取指标

        Args:
            db: 数据库会话
            indicator_id: 指标ID
            load_relations: 是否加载关联的资产

        Returns:
            指标对象，不存在返回None
        """
        try:
            query = select(IndicatorSystem).where(IndicatorSystem.id == indicator_id)

            if load_relations:
                query = query.options(
                    selectinload(IndicatorSystem.asset_relations)
                )

            result = await db.execute(query)
            return result.scalar_one_or_none()

        except Exception as e:
            logger.error(f"获取指标失败: {e}")
            return None

    async def update_indicator(
            self,
            db: AsyncSession,
            indicator_id: int,
            indicator_data: IndicatorSystemUpdate,
            updater: Optional[str] = None
    ) -> IndicatorSystem:
        """
        更新指标

        Args:
            db: 数据库会话
            indicator_id: 指标ID
            indicator_data: 更新数据
            updater: 更新人

        Returns:
            更新后的指标对象

        Raises:
            ValueError: 指标不存在
        """
        try:
            # 获取指标
            indicator = await self.get_indicator_by_id(db, indicator_id)
            if not indicator:
                raise ValueError(f"指标 ID {indicator_id} 不存在")

            # 更新字段
            update_data = indicator_data.model_dump(exclude_unset=True)
            for field, value in update_data.items():
                setattr(indicator, field, value)

            indicator.updated_by = updater

            await db.commit()
            await db.refresh(indicator)

            logger.info(f"更新指标成功: {indicator.indicator_name} (ID: {indicator.id})")
            return indicator

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"更新指标失败: {e}")
            raise ValueError(f"更新指标失败: {str(e)}")

    async def delete_indicator(
            self,
            db: AsyncSession,
            indicator_id: int,
            force: bool = False
    ) -> bool:
        """
        删除指标

        Args:
            db: 数据库会话
            indicator_id: 指标ID
            force: 是否强制删除（物理删除），否则软删除

        Returns:
            是否删除成功

        Raises:
            ValueError: 指标不存在
        """
        try:
            # 获取指标
            indicator = await self.get_indicator_by_id(db, indicator_id)
            if not indicator:
                raise ValueError(f"指标 ID {indicator_id} 不存在")

            # 删除指标
            if force:
                # 物理删除（会级联删除关联关系）
                await db.delete(indicator)
            else:
                # 软删除（修改状态）
                indicator.is_active = False

            await db.commit()

            logger.info(f"删除指标成功: {indicator.indicator_name} (ID: {indicator.id}, force={force})")
            return True

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"删除指标失败: {e}")
            raise ValueError(f"删除指标失败: {str(e)}")

    async def search_indicators(
            self,
            db: AsyncSession,
            page: int = 1,
            page_size: int = 20,
            keyword: Optional[str] = None,
            business_domain: Optional[str] = None,
            business_theme: Optional[str] = None,
            indicator_category: Optional[str] = None,
            indicator_type: Optional[str] = None,
            responsible_dept: Optional[str] = None,
            collection_frequency: Optional[str] = None,
            is_active: Optional[bool] = None
    ) -> Tuple[List[IndicatorSystem], int]:
        """搜索指标（分页）"""
        try:
            # 构建查询条件
            conditions = []

            if keyword:
                keyword_pattern = f"%{keyword}%"
                conditions.append(
                    or_(
                        IndicatorSystem.indicator_name.like(keyword_pattern),
                        IndicatorSystem.indicator_description.like(keyword_pattern),
                        IndicatorSystem.source_system.like(keyword_pattern)
                    )
                )

            if business_domain:
                conditions.append(IndicatorSystem.business_domain == business_domain)

            if business_theme:
                conditions.append(IndicatorSystem.business_theme == business_theme)

            if indicator_category:
                conditions.append(IndicatorSystem.indicator_category == indicator_category)

            if indicator_type:
                conditions.append(IndicatorSystem.indicator_type == indicator_type)

            if responsible_dept:
                conditions.append(IndicatorSystem.responsible_dept == responsible_dept)

            if collection_frequency:
                conditions.append(IndicatorSystem.collection_frequency == collection_frequency)

            if is_active is not None:
                conditions.append(IndicatorSystem.is_active == is_active)

            # 查询总数
            count_query = select(func.count(IndicatorSystem.id))
            if conditions:
                count_query = count_query.where(and_(*conditions))

            count_result = await db.execute(count_query)
            total = count_result.scalar()

            # 查询数据
            query = select(IndicatorSystem)
            if conditions:
                query = query.where(and_(*conditions))

            # 🆕 添加预加载
            query = query.options(selectinload(IndicatorSystem.asset_relations))

            query = query.order_by(desc(IndicatorSystem.created_at))
            query = query.offset((page - 1) * page_size).limit(page_size)

            result = await db.execute(query)
            indicators = result.scalars().all()

            return list(indicators), total

        except Exception as e:
            logger.error(f"搜索指标失败: {e}")
            return [], 0

    async def link_assets(
            self,
            db: AsyncSession,
            indicator_id: int,
            asset_ids: List[int],
            relation_type: str = "source",
            relation_description: Optional[str] = None,
            creator: Optional[str] = None
    ) -> List[IndicatorAssetRelation]:
        """
        关联数据资产到指标

        Args:
            db: 数据库会话
            indicator_id: 指标ID
            asset_ids: 资产ID列表
            relation_type: 关联类型
            relation_description: 关联说明
            creator: 创建人

        Returns:
            创建的关联关系列表

        Raises:
            ValueError: 指标不存在或资产不存在
        """
        try:
            # 验证指标是否存在
            indicator = await self.get_indicator_by_id(db, indicator_id)
            if not indicator:
                raise ValueError(f"指标 ID {indicator_id} 不存在")

            # 验证资产是否存在
            assets_query = select(DataAsset).where(DataAsset.id.in_(asset_ids))
            assets_result = await db.execute(assets_query)
            assets = assets_result.scalars().all()

            if len(assets) != len(asset_ids):
                raise ValueError("部分数据资产不存在")

            # 创建关联关系
            relations = []
            for i, asset_id in enumerate(asset_ids):
                # 检查是否已存在关联
                existing_query = select(IndicatorAssetRelation).where(
                    and_(
                        IndicatorAssetRelation.indicator_id == indicator_id,
                        IndicatorAssetRelation.asset_id == asset_id
                    )
                )
                existing_result = await db.execute(existing_query)
                existing = existing_result.scalar_one_or_none()

                if existing:
                    logger.warning(f"关联已存在: 指标{indicator_id} - 资产{asset_id}")
                    continue

                relation = IndicatorAssetRelation(
                    indicator_id=indicator_id,
                    asset_id=asset_id,
                    relation_type=relation_type,
                    relation_description=relation_description,
                    sort_order=i,
                    created_by=creator
                )
                db.add(relation)
                relations.append(relation)

            await db.commit()

            logger.info(f"关联资产成功: 指标{indicator_id} 关联了 {len(relations)} 个资产")
            return relations

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"关联资产失败: {e}")
            raise ValueError(f"关联资产失败: {str(e)}")

    async def unlink_asset(
            self,
            db: AsyncSession,
            indicator_id: int,
            asset_id: int
    ) -> bool:
        """
        解除指标与资产的关联

        Args:
            db: 数据库会话
            indicator_id: 指标ID
            asset_id: 资产ID

        Returns:
            是否解除成功
        """
        try:
            query = select(IndicatorAssetRelation).where(
                and_(
                    IndicatorAssetRelation.indicator_id == indicator_id,
                    IndicatorAssetRelation.asset_id == asset_id
                )
            )
            result = await db.execute(query)
            relation = result.scalar_one_or_none()

            if not relation:
                raise ValueError(f"关联关系不存在")

            await db.delete(relation)
            await db.commit()

            logger.info(f"解除关联成功: 指标{indicator_id} - 资产{asset_id}")
            return True

        except Exception as e:
            await db.rollback()
            logger.error(f"解除关联失败: {e}")
            raise ValueError(f"解除关联失败: {str(e)}")

    async def get_linked_assets(
            self,
            db: AsyncSession,
            indicator_id: int
    ) -> List[Dict[str, Any]]:
        """
        获取指标关联的所有资产

        Args:
            db: 数据库会话
            indicator_id: 指标ID

        Returns:
            关联的资产列表
        """
        try:
            query = select(IndicatorAssetRelation).where(
                IndicatorAssetRelation.indicator_id == indicator_id
            ).options(
                selectinload(IndicatorAssetRelation.asset)
            ).order_by(IndicatorAssetRelation.sort_order)

            result = await db.execute(query)
            relations = result.scalars().all()

            assets_list = []
            for relation in relations:
                if relation.asset:
                    assets_list.append({
                        "relation_id": relation.id,
                        "asset_id": relation.asset_id,
                        "asset_name": relation.asset.asset_name,
                        "asset_code": relation.asset.asset_code,
                        "table_name": relation.asset.table_name,
                        "relation_type": relation.relation_type,
                        "relation_description": relation.relation_description,
                        "sort_order": relation.sort_order
                    })

            return assets_list

        except Exception as e:
            logger.error(f"获取关联资产失败: {e}")
            return []

    async def batch_create_indicators(
            self,
            db: AsyncSession,
            indicators_data: List[IndicatorSystemCreate],
            creator: Optional[str] = None
    ) -> Tuple[List[IndicatorSystem], List[Dict[str, Any]]]:
        """
        批量创建指标

        Args:
            db: 数据库会话
            indicators_data: 指标创建数据列表
            creator: 创建人

        Returns:
            (成功创建的指标列表, 失败的项目列表)
        """
        success_indicators = []
        failed_items = []

        for idx, indicator_data in enumerate(indicators_data):
            try:
                indicator = await self.create_indicator(db, indicator_data, creator)
                success_indicators.append(indicator)
            except Exception as e:
                failed_items.append({
                    "row": idx + 1,
                    "indicator_name": indicator_data.indicator_name,
                    "error": str(e)
                })
                logger.warning(f"批量创建第 {idx + 1} 条失败: {e}")

        return success_indicators, failed_items

    def generate_excel_template(self) -> BytesIO:
        """
        生成Excel导入模板

        Returns:
            Excel文件的BytesIO对象
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "指标体系模板"

            # 样式定义
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 第一行：中文表头
            for col_idx, header in enumerate(self.EXCEL_HEADERS_CN, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # 第二行：英文字段名（灰色背景）
            field_font = Font(name='Consolas', size=9, color='666666')
            field_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

            for col_idx, field in enumerate(self.EXCEL_HEADERS_EN, start=1):
                cell = ws.cell(row=2, column=col_idx)
                cell.value = field
                cell.font = field_font
                cell.fill = field_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border

            # 第三行：示例数据
            example_data = [
                "CRM系统", "客户管理", "客户分析", "基础指标", "客户总数",
                "统计所有注册客户的总数量", "核心指标",
                "累计值", "数值类", "INTEGER", "10", "整数",
                "市场部", "每日", "23:59", "完全共享", "对外开放"
            ]

            for col_idx, value in enumerate(example_data, start=1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = value
                cell.border = border

            # 设置列宽
            column_widths = [15, 15, 15, 15, 25, 30, 20, 15, 20, 15, 12, 15, 15, 15, 15, 15, 15]
            for idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

            # 冻结前两行
            ws.freeze_panes = 'A3'

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info("生成Excel模板成功")
            return output

        except Exception as e:
            logger.error(f"生成Excel模板失败: {e}")
            raise ValueError(f"生成Excel模板失败: {str(e)}")

    async def import_from_excel(
            self,
            db: AsyncSession,
            file_content: bytes,
            creator: Optional[str] = None
    ) -> Tuple[List[IndicatorSystem], List[Dict[str, Any]]]:
        """从Excel导入指标（带重复检测）"""
        success_indicators = []
        failed_rows = []
        BATCH_SIZE = 100

        try:
            wb = openpyxl.load_workbook(BytesIO(file_content))
            ws = wb.active

            indicators_batch = []

            # 🆕 用于检测重复的集合
            seen_indicators = set()

            # 🆕 从数据库加载已存在的指标名称
            existing_query = select(IndicatorSystem.indicator_name).where(
                IndicatorSystem.is_active == True
            )
            existing_result = await db.execute(existing_query)
            existing_names = {row[0] for row in existing_result}

            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    if not any(row):
                        continue

                    indicator_name = row[4] if row[4] else f"未命名指标_{row_idx}"

                    # 🆕 检查Excel内部重复
                    if indicator_name in seen_indicators:
                        failed_rows.append({
                            "row": row_idx,
                            "data": list(row) if row else [],
                            "error": f"指标名称重复（Excel内）: {indicator_name}"
                        })
                        continue

                    # 🆕 检查数据库中已存在
                    if indicator_name in existing_names:
                        failed_rows.append({
                            "row": row_idx,
                            "data": list(row) if row else [],
                            "error": f"指标名称已存在于数据库: {indicator_name}"
                        })
                        continue

                    seen_indicators.add(indicator_name)

                    indicator = IndicatorSystem(
                        source_system=row[0] if row[0] else None,
                        business_domain=row[1] if row[1] else None,
                        business_theme=row[2] if row[2] else None,
                        indicator_category=row[3] if row[3] else None,
                        indicator_name=indicator_name,
                        indicator_description=row[5] if row[5] else None,
                        remark=row[6] if row[6] else None,
                        indicator_type=row[7] if row[7] else None,
                        tech_classification=row[8] if row[8] else None,
                        data_type=row[9] if row[9] else None,
                        data_length=int(row[10]) if row[10] and str(row[10]).replace('.', '').replace('-',
                                                                                                      '').isdigit() else None,
                        data_format=row[11] if row[11] else None,
                        responsible_dept=row[12] if row[12] else None,
                        collection_frequency=row[13] if row[13] else None,
                        collection_time=row[14] if row[14] else None,
                        share_type=row[15] if row[15] else None,
                        open_attribute=row[16] if row[16] else None,
                        is_active=True,
                        created_by=creator,
                        updated_by=creator
                    )

                    indicators_batch.append(indicator)

                    # 每100条提交一次
                    if len(indicators_batch) >= BATCH_SIZE:
                        try:
                            db.add_all(indicators_batch)
                            await db.commit()

                            for ind in indicators_batch:
                                await db.refresh(ind)
                                success_indicators.append(ind)
                                existing_names.add(ind.indicator_name)  # 🆕 更新已存在集合

                            logger.info(f"已导入 {len(success_indicators)} 条")
                            indicators_batch = []

                        except Exception as e:
                            await db.rollback()
                            logger.error(f"批次导入失败: {e}")
                            for ind in indicators_batch:
                                failed_rows.append({
                                    "row": row_idx,
                                    "data": [ind.indicator_name],
                                    "error": str(e)
                                })
                            indicators_batch = []

                except Exception as e:
                    failed_rows.append({
                        "row": row_idx,
                        "data": list(row) if row else [],
                        "error": str(e)
                    })

            # 提交剩余的数据
            if indicators_batch:
                try:
                    db.add_all(indicators_batch)
                    await db.commit()

                    for ind in indicators_batch:
                        await db.refresh(ind)
                        success_indicators.append(ind)

                except Exception as e:
                    await db.rollback()
                    logger.error(f"最后批次导入失败: {e}")
                    for ind in indicators_batch:
                        failed_rows.append({
                            "row": -1,
                            "data": [ind.indicator_name],
                            "error": str(e)
                        })

            logger.info(f"Excel导入完成: 成功 {len(success_indicators)}，失败 {len(failed_rows)}")
            return success_indicators, failed_rows

        except Exception as e:
            logger.error(f"解析Excel失败: {e}")
            raise ValueError(f"解析Excel失败: {str(e)}")

    def export_to_excel(self, indicators: List[IndicatorSystem]) -> BytesIO:
        """
        将指标数据导出为Excel
        """
        try:
            if indicators:
                first = indicators[0]
                logger.info(f"第一条数据类型检查:")
                logger.info(f"source_system: {type(first.source_system)} = {first.source_system}")
                logger.info(f"business_domain: {type(first.business_domain)} = {first.business_domain}")
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "指标数据"

            # 样式定义
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 写入表头
            for col_idx, header in enumerate(self.EXCEL_HEADERS_CN, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # 写入数据
            for row_idx, indicator in enumerate(indicators, start=2):
                data_row = [
                    str(indicator.source_system) if indicator.source_system else '',
                    str(indicator.business_domain) if indicator.business_domain else '',
                    str(indicator.business_theme) if indicator.business_theme else '',
                    str(indicator.indicator_category) if indicator.indicator_category else '',
                    str(indicator.indicator_name) if indicator.indicator_name else '',
                    str(indicator.indicator_description) if indicator.indicator_description else '',
                    str(indicator.remark) if indicator.remark else '',
                    str(indicator.indicator_type) if indicator.indicator_type else '',
                    str(indicator.tech_classification) if indicator.tech_classification else '',
                    str(indicator.data_type) if indicator.data_type else '',
                    str(indicator.data_length) if indicator.data_length is not None else '',
                    str(indicator.data_format) if indicator.data_format else '',
                    str(indicator.responsible_dept) if indicator.responsible_dept else '',
                    str(indicator.collection_frequency) if indicator.collection_frequency else '',
                    str(indicator.collection_time) if indicator.collection_time else '',
                    str(indicator.share_type) if indicator.share_type else '',
                    str(indicator.open_attribute) if indicator.open_attribute else ''
                ]

                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value  # value已经是字符串了
                    cell.border = border

            # 设置列宽
            column_widths = [15, 15, 15, 20, 25, 30, 20, 15, 20, 15, 12, 15, 15, 15, 25, 15, 15]
            for idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info(f"导出Excel成功，共 {len(indicators)} 条数据")
            return output

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise ValueError(f"导出Excel失败: {str(e)}")


    async def get_statistics(self, db: AsyncSession) -> IndicatorStatistics:
        """
        获取指标统计信息

        Args:
            db: 数据库会话

        Returns:
            统计信息
        """
        try:
            # 总数
            total_query = select(func.count(IndicatorSystem.id))
            total_result = await db.execute(total_query)
            total_count = total_result.scalar()

            # 启用数
            active_query = select(func.count(IndicatorSystem.id)).where(
                IndicatorSystem.is_active == True
            )
            active_result = await db.execute(active_query)
            active_count = active_result.scalar()

            # 按业务领域统计
            domain_query = select(
                IndicatorSystem.business_domain,
                func.count(IndicatorSystem.id)
            ).where(
                IndicatorSystem.business_domain.isnot(None)
            ).group_by(IndicatorSystem.business_domain)
            domain_result = await db.execute(domain_query)
            by_domain = {row[0]: row[1] for row in domain_result}

            # 按指标类别统计
            category_query = select(
                IndicatorSystem.indicator_category,
                func.count(IndicatorSystem.id)
            ).where(
                IndicatorSystem.indicator_category.isnot(None)
            ).group_by(IndicatorSystem.indicator_category)
            category_result = await db.execute(category_query)
            by_category = {row[0]: row[1] for row in category_result}

            # 按采集频率统计
            frequency_query = select(
                IndicatorSystem.collection_frequency,
                func.count(IndicatorSystem.id)
            ).where(
                IndicatorSystem.collection_frequency.isnot(None)
            ).group_by(IndicatorSystem.collection_frequency)
            frequency_result = await db.execute(frequency_query)
            by_frequency = {row[0]: row[1] for row in frequency_result}

            return IndicatorStatistics(
                total_count=total_count,
                active_count=active_count,
                inactive_count=total_count - active_count,
                by_domain=by_domain,
                by_category=by_category,
                by_frequency=by_frequency
            )

        except Exception as e:
            logger.error(f"获取统计信息失败: {e}")
            return IndicatorStatistics(
                total_count=0,
                active_count=0,
                inactive_count=0,
                by_domain={},
                by_category={},
                by_frequency={}
            )

    async def get_all_business_domains(self, db: AsyncSession) -> List[str]:
        """获取所有业务领域（去重）"""
        try:
            query = select(IndicatorSystem.business_domain).where(
                IndicatorSystem.business_domain.isnot(None)
            ).distinct().order_by(IndicatorSystem.business_domain)

            result = await db.execute(query)
            return [row[0] for row in result]
        except Exception as e:
            logger.error(f"获取业务领域列表失败: {e}")
            return []

    async def get_all_indicator_categories(self, db: AsyncSession) -> List[str]:
        """获取所有指标类别（去重）"""
        try:
            query = select(IndicatorSystem.indicator_category).where(
                IndicatorSystem.indicator_category.isnot(None)
            ).distinct().order_by(IndicatorSystem.indicator_category)

            result = await db.execute(query)
            return [row[0] for row in result]
        except Exception as e:
            logger.error(f"获取指标类别列表失败: {e}")
            return []


# 创建全局服务实例
indicator_system_service = IndicatorSystemService()