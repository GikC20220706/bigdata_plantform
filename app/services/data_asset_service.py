# app/services/data_asset_service.py
"""
数据资产服务层
处理数据资产的业务逻辑
"""
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
from sqlalchemy import select, and_, or_, func, desc
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload, joinedload
from loguru import logger
from datetime import datetime

from app.models.data_asset import DataAsset, AssetColumn, AssetAccessLog
from app.models.data_catalog import DataCatalog
from app.models.data_source import DataSource
from app.schemas.data_catalog import (
    DataAssetCreate, DataAssetUpdate,
    DataAssetResponse, AssetColumnDetail
)
from app.services.data_catalog_service import data_catalog_service


class DataAssetService:
    """数据资产服务类"""

    async def create_asset(
            self,
            db: AsyncSession,
            asset_data: DataAssetCreate,
            creator: Optional[str] = None,
            auto_fetch_metadata: bool = True
    ) -> DataAsset:
        """
        创建数据资产

        Args:
            db: 数据库会话
            asset_data: 资产创建数据
            creator: 创建人
            auto_fetch_metadata: 是否自动获取表元数据

        Returns:
            创建的资产对象

        Raises:
            ValueError: 编码重复、目录不存在、数据源不存在等
        """
        try:
            # 1. 检查编码是否重复
            existing = await self._get_by_code(db, asset_data.asset_code)
            if existing:
                raise ValueError(f"资产编码 {asset_data.asset_code} 已存在")

            # 2. 验证目录是否存在
            catalog = await data_catalog_service.get_catalog_by_id(db, asset_data.catalog_id)
            if not catalog:
                raise ValueError(f"目录 ID {asset_data.catalog_id} 不存在")

            # 3. 验证数据源是否存在
            data_source = await self._get_data_source(db, asset_data.data_source_id)
            if not data_source:
                raise ValueError(f"数据源 ID {asset_data.data_source_id} 不存在")

            # 4. 创建资产对象
            asset = DataAsset(
                asset_name=asset_data.asset_name,
                asset_code=asset_data.asset_code,
                asset_type=asset_data.asset_type.value,
                catalog_id=asset_data.catalog_id,
                data_source_id=asset_data.data_source_id,
                database_name=asset_data.database_name,
                schema_name=asset_data.schema_name,
                table_name=asset_data.table_name,
                business_description=asset_data.business_description,
                technical_description=asset_data.technical_description,
                quality_level=asset_data.quality_level.value,
                usage_frequency=asset_data.usage_frequency,
                importance_level=asset_data.importance_level,
                status="normal",
                is_public=True,
                is_api_published=False,
                preview_count=0,
                download_count=0,
                api_call_count=0,
                allow_download=asset_data.allow_download,
                max_download_rows=asset_data.max_download_rows,
                tags=asset_data.tags,
                business_owner=asset_data.business_owner,
                technical_owner=asset_data.technical_owner,
                created_by=creator,
                updated_by=creator
            )

            db.add(asset)
            await db.flush()  # 先flush获取ID，但不提交

            # 5. 如果需要自动获取元数据
            if auto_fetch_metadata:
                try:
                    await self._fetch_and_save_metadata(db, asset, data_source)
                except Exception as e:
                    logger.warning(f"获取表元数据失败，将继续创建资产: {e}")

            # 6. 更新目录的资产计数
            await data_catalog_service.update_asset_count(db, catalog.id, increment=1)

            await db.commit()
            await db.refresh(asset)

            logger.info(f"创建数据资产成功: {asset.asset_name} (ID: {asset.id})")
            return asset

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"创建数据资产失败: {e}")
            raise ValueError(f"创建资产失败: {str(e)}")

    async def get_asset_by_id(
            self,
            db: AsyncSession,
            asset_id: int,
            load_relationships: bool = False
    ) -> Optional[DataAsset]:
        """
        根据ID获取资产

        Args:
            db: 数据库会话
            asset_id: 资产ID
            load_relationships: 是否加载关联对象（目录、数据源等）

        Returns:
            资产对象，不存在返回None
        """
        try:
            query = select(DataAsset).where(DataAsset.id == asset_id)

            if load_relationships:
                query = query.options(
                    selectinload(DataAsset.catalog),
                    selectinload(DataAsset.data_source),
                    selectinload(DataAsset.columns)
                )

            result = await db.execute(query)
            return result.scalar_one_or_none()

        except Exception as e:
            logger.error(f"获取资产失败: {e}")
            return None

    async def update_asset(
            self,
            db: AsyncSession,
            asset_id: int,
            asset_data: DataAssetUpdate,
            updater: Optional[str] = None
    ) -> DataAsset:
        """
        更新数据资产

        Args:
            db: 数据库会话
            asset_id: 资产ID
            asset_data: 更新数据
            updater: 更新人

        Returns:
            更新后的资产对象

        Raises:
            ValueError: 资产不存在
        """
        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 记录原目录ID（用于更新计数）
            old_catalog_id = asset.catalog_id

            # 3. 更新字段
            update_data = asset_data.model_dump(exclude_unset=True)
            for field, value in update_data.items():
                if field == "quality_level" and value:
                    setattr(asset, field, value.value)
                elif field == "status" and value:
                    setattr(asset, field, value.value)
                else:
                    setattr(asset, field, value)

            asset.updated_by = updater

            # 4. 如果目录变了，需要更新目录的资产计数
            if asset_data.catalog_id and asset_data.catalog_id != old_catalog_id:
                # 旧目录减1
                await data_catalog_service.update_asset_count(db, old_catalog_id, increment=-1)
                # 新目录加1
                await data_catalog_service.update_asset_count(db, asset_data.catalog_id, increment=1)

            await db.commit()
            await db.refresh(asset)

            logger.info(f"更新数据资产成功: {asset.asset_name} (ID: {asset.id})")
            return asset

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"更新数据资产失败: {e}")
            raise ValueError(f"更新资产失败: {str(e)}")

    async def delete_asset(
            self,
            db: AsyncSession,
            asset_id: int,
            force: bool = False
    ) -> bool:
        """
        删除数据资产

        Args:
            db: 数据库会话
            asset_id: 资产ID
            force: 是否强制删除（物理删除），否则软删除

        Returns:
            是否删除成功

        Raises:
            ValueError: 资产不存在
        """
        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 检查是否已发布API
            if not force and asset.is_api_published:
                raise ValueError("资产已发布为API，请先取消发布后再删除")

            # 3. 删除资产
            if force:
                # 物理删除
                await db.delete(asset)
            else:
                # 软删除（修改状态）
                asset.status = "deprecated"
                asset.is_public = False

            # 4. 更新目录的资产计数
            await data_catalog_service.update_asset_count(db, asset.catalog_id, increment=-1)

            await db.commit()

            logger.info(f"删除数据资产成功: {asset.asset_name} (ID: {asset.id}, force={force})")
            return True

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"删除数据资产失败: {e}")
            raise ValueError(f"删除资产失败: {str(e)}")

    async def search_assets(
            self,
            db: AsyncSession,
            page: int = 1,
            page_size: int = 20,
            keyword: Optional[str] = None,
            catalog_id: Optional[int] = None,
            data_source_id: Optional[int] = None,
            asset_type: Optional[str] = None,
            status: Optional[str] = None,
            quality_level: Optional[str] = None
    ) -> Tuple[List[DataAsset], int]:
        """
        搜索数据资产（分页）

        Args:
            db: 数据库会话
            page: 页码
            page_size: 每页数量
            keyword: 搜索关键词（名称、编码、描述）
            catalog_id: 目录ID
            data_source_id: 数据源ID
            asset_type: 资产类型
            status: 状态
            quality_level: 质量等级

        Returns:
            (资产列表, 总数)
        """
        try:
            # 1. 构建查询条件
            conditions = []

            if keyword:
                keyword_pattern = f"%{keyword}%"
                conditions.append(
                    or_(
                        DataAsset.asset_name.like(keyword_pattern),
                        DataAsset.asset_code.like(keyword_pattern),
                        DataAsset.table_name.like(keyword_pattern),
                        DataAsset.business_description.like(keyword_pattern)
                    )
                )

            if catalog_id:
                conditions.append(DataAsset.catalog_id == catalog_id)

            if data_source_id:
                conditions.append(DataAsset.data_source_id == data_source_id)

            if asset_type:
                conditions.append(DataAsset.asset_type == asset_type)

            if status:
                conditions.append(DataAsset.status == status)

            if quality_level:
                conditions.append(DataAsset.quality_level == quality_level)

            # 2. 查询总数
            count_query = select(func.count(DataAsset.id))
            if conditions:
                count_query = count_query.where(and_(*conditions))

            total_result = await db.execute(count_query)
            total = total_result.scalar_one()

            # 3. 分页查询（预加载关联对象）
            query = (
                select(DataAsset)
                .options(
                    selectinload(DataAsset.catalog),
                    selectinload(DataAsset.data_source)
                )
            )

            if conditions:
                query = query.where(and_(*conditions))

            query = query.order_by(
                desc(DataAsset.created_at)
            ).limit(page_size).offset((page - 1) * page_size)

            result = await db.execute(query)
            assets = result.scalars().all()

            return list(assets), total

        except Exception as e:
            logger.error(f"搜索资产失败: {e}")
            raise ValueError(f"搜索资产失败: {str(e)}")

    async def get_asset_columns(
            self,
            db: AsyncSession,
            asset_id: int
    ) -> List[AssetColumn]:
        """
        获取资产的字段列表

        Args:
            db: 数据库会话
            asset_id: 资产ID

        Returns:
            字段列表
        """
        try:
            result = await db.execute(
                select(AssetColumn)
                .where(AssetColumn.asset_id == asset_id)
                .order_by(AssetColumn.sort_order, AssetColumn.id)
            )
            return list(result.scalars().all())

        except Exception as e:
            logger.error(f"获取资产字段失败: {e}")
            return []

    async def refresh_asset_metadata(
            self,
            db: AsyncSession,
            asset_id: int
    ) -> DataAsset:
        """
        刷新资产元数据（从数据源重新获取）

        Args:
            db: 数据库会话
            asset_id: 资产ID

        Returns:
            更新后的资产对象

        Raises:
            ValueError: 资产不存在或获取元数据失败
        """
        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id, load_relationships=True)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 获取数据源
            data_source = asset.data_source
            if not data_source:
                raise ValueError("资产关联的数据源不存在")

            # 3. 删除旧的字段信息
            await db.execute(
                select(AssetColumn).where(AssetColumn.asset_id == asset_id)
            )
            old_columns = await self.get_asset_columns(db, asset_id)
            for col in old_columns:
                await db.delete(col)

            # 4. 重新获取元数据
            await self._fetch_and_save_metadata(db, asset, data_source)

            # 5. 更新资产的最后更新时间
            asset.last_updated_at = datetime.now()

            await db.commit()
            await db.refresh(asset)

            logger.info(f"刷新资产元数据成功: {asset.asset_name} (ID: {asset.id})")
            return asset

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"刷新资产元数据失败: {e}")
            raise ValueError(f"刷新元数据失败: {str(e)}")

    async def batch_create_assets(
            self,
            db: AsyncSession,
            assets_data: List[DataAssetCreate],
            creator: Optional[str] = None
    ) -> Tuple[List[DataAsset], List[Dict[str, Any]]]:
        """
        批量创建资产

        Args:
            db: 数据库会话
            assets_data: 资产创建数据列表
            creator: 创建人

        Returns:
            (成功创建的资产列表, 失败的项目列表)
        """
        success_assets = []
        failed_items = []

        for idx, asset_data in enumerate(assets_data):
            try:
                asset = await self.create_asset(
                    db,
                    asset_data,
                    creator,
                    auto_fetch_metadata=False  # 批量创建时不自动获取元数据，提高速度
                )
                success_assets.append(asset)

            except Exception as e:
                logger.error(f"批量创建第 {idx + 1} 个资产失败: {e}")
                failed_items.append({
                    "index": idx + 1,
                    "asset_name": asset_data.asset_name,
                    "error": str(e)
                })

        logger.info(
            f"批量创建资产完成: 成功 {len(success_assets)}，"
            f"失败 {len(failed_items)}"
        )

        return success_assets, failed_items

    async def preview_asset_data(
            self,
            db: AsyncSession,
            asset_id: int,
            limit: int = 100,
            offset: int = 0,
            columns: Optional[List[str]] = None,
            filter_conditions: Optional[Dict[str, Any]] = None,
            user: Optional[str] = None,
            user_ip: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        预览资产数据

        Args:
            db: 数据库会话
            asset_id: 资产ID
            limit: 预览行数
            offset: 偏移量
            columns: 指定列，None表示全部
            filter_conditions: 过滤条件
            user: 访问用户
            user_ip: 访问IP

        Returns:
            预览结果字典

        Raises:
            ValueError: 资产不存在或预览失败
        """
        start_time = datetime.now()

        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id, load_relationships=True)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 获取数据源
            data_source = asset.data_source
            if not data_source:
                raise ValueError("资产关联的数据源不存在")

            # 3. 构建查询
            from app.services.optimized_data_integration_service import get_optimized_data_integration_service
            integration_service = get_optimized_data_integration_service()

            # 构建SQL查询
            select_columns = columns if columns else ["*"]
            sql = f"SELECT {', '.join(select_columns)} FROM "

            # 添加数据库和表名
            if asset.database_name:
                sql += f"{asset.database_name}."
            if asset.schema_name:
                sql += f"{asset.schema_name}."
            sql += asset.table_name

            # 添加过滤条件（简单实现）
            if filter_conditions:
                where_clauses = []
                for key, value in filter_conditions.items():
                    if isinstance(value, str):
                        where_clauses.append(f"{key} = '{value}'")
                    else:
                        where_clauses.append(f"{key} = {value}")
                if where_clauses:
                    sql += f" WHERE {' AND '.join(where_clauses)}"

            # 添加分页
            sql += f" LIMIT {limit} OFFSET {offset}"

            logger.info(f"预览SQL: {sql}")

            # 4. 执行查询
            result = await integration_service.execute_query(
                data_source.name,
                sql,
                database=asset.database_name,
                limit=limit
            )

            if not result.get('success'):
                raise ValueError(f"查询失败: {result.get('error')}")

            data_rows = result.get('data', [])

            # 5. 获取列信息
            column_info = []
            if data_rows:
                for col_name in data_rows[0].keys():
                    column_info.append({
                        'name': col_name,
                        'type': 'string'  # 简化处理
                    })

            # 6. 记录访问日志
            response_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            await self._log_access(
                db, asset_id, "preview", user, user_ip,
                len(data_rows), None, "success", None, response_time_ms
            )

            # 7. 更新预览次数
            asset.preview_count += 1
            asset.last_accessed_at = datetime.now()
            await db.commit()

            return {
                'asset_id': asset_id,
                'asset_name': asset.asset_name,
                'columns': column_info,
                'data': data_rows,
                'total_rows': None,  # 可选：执行COUNT查询获取总数
                'preview_rows': len(data_rows),
                'has_more': len(data_rows) == limit,
                'preview_time': datetime.now()
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"预览资产数据失败: {e}")

            # 记录失败日志
            response_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            await self._log_access(
                db, asset_id, "preview", user, user_ip,
                0, None, "failed", str(e), response_time_ms
            )

            raise ValueError(f"预览数据失败: {str(e)}")

    async def download_asset_data(
            self,
            db: AsyncSession,
            asset_id: int,
            format: str = "excel",
            columns: Optional[List[str]] = None,
            filter_conditions: Optional[Dict[str, Any]] = None,
            max_rows: Optional[int] = None,
            filename: Optional[str] = None,
            user: Optional[str] = None,
            user_ip: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        下载资产数据

        Args:
            db: 数据库会话
            asset_id: 资产ID
            format: 下载格式 (excel/csv/json)
            columns: 指定列
            filter_conditions: 过滤条件
            max_rows: 最大行数
            filename: 文件名
            user: 访问用户
            user_ip: 访问IP

        Returns:
            下载信息字典

        Raises:
            ValueError: 资产不存在、不允许下载等
        """
        import uuid
        import os
        from pathlib import Path

        start_time = datetime.now()

        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id, load_relationships=True)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 检查是否允许下载
            if not asset.allow_download:
                raise ValueError("该资产不允许下载")

            # 3. 确定最大行数
            actual_max_rows = max_rows or asset.max_download_rows
            if actual_max_rows > asset.max_download_rows:
                actual_max_rows = asset.max_download_rows
                logger.warning(f"下载行数超过限制，已调整为 {actual_max_rows}")

            # 4. 获取数据
            data_source = asset.data_source
            if not data_source:
                raise ValueError("资产关联的数据源不存在")

            from app.services.optimized_data_integration_service import get_optimized_data_integration_service
            integration_service = get_optimized_data_integration_service()

            # 构建SQL查询
            select_columns = columns if columns else ["*"]
            sql = f"SELECT {', '.join(select_columns)} FROM "

            if asset.database_name:
                sql += f"{asset.database_name}."
            if asset.schema_name:
                sql += f"{asset.schema_name}."
            sql += asset.table_name

            # 添加过滤条件
            if filter_conditions:
                where_clauses = []
                for key, value in filter_conditions.items():
                    if isinstance(value, str):
                        where_clauses.append(f"{key} = '{value}'")
                    else:
                        where_clauses.append(f"{key} = {value}")
                if where_clauses:
                    sql += f" WHERE {' AND '.join(where_clauses)}"

            sql += f" LIMIT {actual_max_rows}"

            logger.info(f"下载SQL: {sql}")

            # 执行查询
            result = await integration_service.execute_query(
                data_source.name,
                sql,
                database=asset.database_name,
                limit=actual_max_rows
            )

            if not result.get('success'):
                raise ValueError(f"查询失败: {result.get('error')}")

            data_rows = result.get('data', [])

            # 5. 生成文件
            download_id = str(uuid.uuid4())
            file_ext = self._get_file_extension(format)
            actual_filename = filename or f"{asset.asset_code}_{download_id[:8]}{file_ext}"

            # 创建临时下载目录
            download_dir = Path("downloads/temp")
            download_dir.mkdir(parents=True, exist_ok=True)

            file_path = download_dir / actual_filename

            # 根据格式生成文件
            if format == "excel":
                file_size = await self._generate_excel_file(file_path, data_rows)
            elif format == "csv":
                file_size = await self._generate_csv_file(file_path, data_rows)
            elif format == "json":
                file_size = await self._generate_json_file(file_path, data_rows)
            else:
                raise ValueError(f"不支持的格式: {format}")

            # 6. 记录访问日志
            response_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            await self._log_access(
                db, asset_id, "download", user, user_ip,
                len(data_rows), file_size, "success", None, response_time_ms
            )

            # 7. 更新下载次数
            asset.download_count += 1
            asset.last_accessed_at = datetime.now()
            await db.commit()

            # 8. 返回下载信息
            return {
                'download_id': download_id,
                'filename': actual_filename,
                'file_size': file_size,
                'row_count': len(data_rows),
                'format': format,
                'download_url': f"/api/v1/data-asset/{asset_id}/download/{download_id}",
                'expires_at': datetime.now().replace(hour=23, minute=59, second=59)  # 当天过期
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"下载资产数据失败: {e}")

            # 记录失败日志
            response_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            await self._log_access(
                db, asset_id, "download", user, user_ip,
                0, None, "failed", str(e), response_time_ms
            )

            raise ValueError(f"下载数据失败: {str(e)}")

    async def publish_asset_as_api(
            self,
            db: AsyncSession,
            asset_id: int,
            api_config: Dict[str, Any],
            creator: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        将资产发布为API接口

        Args:
            db: 数据库会话
            asset_id: 资产ID
            api_config: API配置（来自PublishAPIRequest）
            creator: 创建人

        Returns:
            发布结果字典

        Raises:
            ValueError: 资产不存在、已发布等
        """
        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id, load_relationships=True)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 检查是否已发布
            if asset.is_api_published:
                raise ValueError("资产已发布为API，请先取消发布")

            # 3. 构建SQL模板
            data_source = asset.data_source
            if not data_source:
                raise ValueError("资产关联的数据源不存在")

            # 构建基础SELECT语句
            select_columns = api_config.get('select_columns', ['*'])
            sql_template = f"SELECT {', '.join(select_columns)} FROM "

            if asset.database_name:
                sql_template += f"{asset.database_name}."
            if asset.schema_name:
                sql_template += f"{asset.schema_name}."
            sql_template += asset.table_name

            # 添加WHERE条件
            where_conditions = api_config.get('where_conditions')
            if where_conditions:
                sql_template += f" WHERE {where_conditions}"

            # 添加ORDER BY
            order_by = api_config.get('order_by')
            if order_by:
                sql_template += f" ORDER BY {order_by}"

            # 添加LIMIT（使用参数化）
            limit = api_config.get('limit', 1000)
            sql_template += f" LIMIT {limit}"

            # 4. 调用CustomAPI服务创建API
            from app.services.custom_api_service import custom_api_service
            from app.schemas.custom_api import CreateAPIRequest, HTTPMethod, ResponseFormat

            # 构建API创建请求
            api_name = api_config.get('api_name')
            api_path = api_config.get('api_path')

            # 确保路径格式正确
            if not api_path.startswith('/api/custom/'):
                api_path = f"/api/custom/{api_path.lstrip('/')}"

            create_api_request = CreateAPIRequest(
                api_name=api_name,
                api_path=api_path,
                description=api_config.get('description', f"从资产 {asset.asset_name} 自动生成的API"),
                data_source_id=asset.data_source_id,
                sql_template=sql_template,
                http_method=HTTPMethod(api_config.get('http_method', 'GET')),
                response_format=ResponseFormat(api_config.get('response_format', 'json')),
                parameters=api_config.get('parameters', []),
                is_active=True,
                is_public=api_config.get('is_public', True),
                rate_limit=api_config.get('rate_limit', 100),
                cache_ttl=api_config.get('cache_ttl', 300)
            )

            # 创建API
            custom_api = await custom_api_service.create_api(db, create_api_request, creator)

            # 5. 更新资产的API发布信息
            asset.is_api_published = True
            asset.published_api_id = custom_api.id
            asset.api_publish_time = datetime.now()

            await db.commit()

            # 6. 注册到动态路由
            from app.utils.dynamic_api_router import dynamic_api_router
            await dynamic_api_router.register_api(custom_api)

            logger.info(f"资产发布为API成功: {asset.asset_name} -> {api_name}")

            # 7. 返回发布信息
            return {
                'api_id': custom_api.id,
                'api_name': custom_api.api_name,
                'api_path': custom_api.api_path,
                'api_url': f"http://localhost:8000{custom_api.api_path}",  # TODO: 使用配置的域名
                'status': 'published',
                'published_at': asset.api_publish_time,
                'test_url': f"http://localhost:8000{custom_api.api_path}?limit=10"
            }

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"发布资产为API失败: {e}")
            raise ValueError(f"发布API失败: {str(e)}")

    async def unpublish_asset_api(
            self,
            db: AsyncSession,
            asset_id: int
    ) -> bool:
        """
        取消发布资产的API

        Args:
            db: 数据库会话
            asset_id: 资产ID

        Returns:
            是否取消成功

        Raises:
            ValueError: 资产不存在或未发布
        """
        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id, load_relationships=True)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 检查是否已发布
            if not asset.is_api_published or not asset.published_api_id:
                raise ValueError("资产未发布API")

            # 3. 删除CustomAPI
            from app.services.custom_api_service import custom_api_service
            await custom_api_service.delete_api(db, asset.published_api_id)

            # 4. 更新资产状态
            asset.is_api_published = False
            asset.published_api_id = None
            asset.api_publish_time = None

            await db.commit()

            logger.info(f"取消资产API发布成功: {asset.asset_name}")
            return True

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"取消API发布失败: {e}")
            raise ValueError(f"取消发布失败: {str(e)}")

    async def get_asset_statistics(
            self,
            db: AsyncSession,
            asset_id: int
    ) -> Dict[str, Any]:
        """
        获取资产统计信息

        Args:
            db: 数据库会话
            asset_id: 资产ID

        Returns:
            统计信息字典
        """
        try:
            # 1. 获取资产
            asset = await self.get_asset_by_id(db, asset_id)
            if not asset:
                raise ValueError(f"资产 ID {asset_id} 不存在")

            # 2. 查询最近的访问日志
            from sqlalchemy import and_

            # 最后预览时间
            preview_result = await db.execute(
                select(AssetAccessLog.access_time)
                .where(
                    and_(
                        AssetAccessLog.asset_id == asset_id,
                        AssetAccessLog.access_type == "preview",
                        AssetAccessLog.status == "success"
                    )
                )
                .order_by(desc(AssetAccessLog.access_time))
                .limit(1)
            )
            last_preview = preview_result.scalar_one_or_none()

            # 最后下载时间
            download_result = await db.execute(
                select(AssetAccessLog.access_time)
                .where(
                    and_(
                        AssetAccessLog.asset_id == asset_id,
                        AssetAccessLog.access_type == "download",
                        AssetAccessLog.status == "success"
                    )
                )
                .order_by(desc(AssetAccessLog.access_time))
                .limit(1)
            )
            last_download = download_result.scalar_one_or_none()

            # 最后API调用时间
            api_result = await db.execute(
                select(AssetAccessLog.access_time)
                .where(
                    and_(
                        AssetAccessLog.asset_id == asset_id,
                        AssetAccessLog.access_type == "api",
                        AssetAccessLog.status == "success"
                    )
                )
                .order_by(desc(AssetAccessLog.access_time))
                .limit(1)
            )
            last_api_call = api_result.scalar_one_or_none()

            # 3. 计算热度评分（简单算法）
            popularity_score = self._calculate_popularity_score(
                asset.preview_count,
                asset.download_count,
                asset.api_call_count
            )

            return {
                'asset_id': asset.id,
                'asset_name': asset.asset_name,
                'total_rows': asset.row_count,
                'total_size': asset.data_size,
                'column_count': asset.column_count or 0,
                'total_preview_count': asset.preview_count,
                'total_download_count': asset.download_count,
                'total_api_call_count': asset.api_call_count,
                'last_preview_time': last_preview,
                'last_download_time': last_download,
                'last_api_call_time': last_api_call,
                'popularity_score': popularity_score
            }

        except Exception as e:
            logger.error(f"获取资产统计失败: {e}")
            raise ValueError(f"获取统计信息失败: {str(e)}")

    async def get_asset_access_logs(
            self,
            db: AsyncSession,
            asset_id: int,
            page: int = 1,
            page_size: int = 20,
            access_type: Optional[str] = None,
            start_time: Optional[datetime] = None,
            end_time: Optional[datetime] = None
    ) -> Tuple[List[AssetAccessLog], int]:
        """
        获取资产访问日志

        Args:
            db: 数据库会话
            asset_id: 资产ID
            page: 页码
            page_size: 每页数量
            access_type: 访问类型筛选
            start_time: 开始时间
            end_time: 结束时间

        Returns:
            (日志列表, 总数)
        """
        try:
            # 构建查询条件
            conditions = [AssetAccessLog.asset_id == asset_id]

            if access_type:
                conditions.append(AssetAccessLog.access_type == access_type)

            if start_time:
                conditions.append(AssetAccessLog.access_time >= start_time)

            if end_time:
                conditions.append(AssetAccessLog.access_time <= end_time)

            # 查询总数
            count_result = await db.execute(
                select(func.count(AssetAccessLog.id))
                .where(and_(*conditions))
            )
            total = count_result.scalar_one()

            # 分页查询
            result = await db.execute(
                select(AssetAccessLog)
                .where(and_(*conditions))
                .order_by(desc(AssetAccessLog.access_time))
                .limit(page_size)
                .offset((page - 1) * page_size)
            )
            logs = result.scalars().all()

            return list(logs), total

        except Exception as e:
            logger.error(f"获取访问日志失败: {e}")
            return [], 0

    # ==================== 辅助私有方法 ====================

    async def _log_access(
            self,
            db: AsyncSession,
            asset_id: int,
            access_type: str,
            access_user: Optional[str],
            access_ip: Optional[str],
            record_count: int,
            file_size: Optional[int],
            status: str,
            error_message: Optional[str],
            response_time_ms: int
    ) -> None:
        """记录访问日志"""
        try:
            log = AssetAccessLog(
                asset_id=asset_id,
                access_type=access_type,
                access_user=access_user,
                access_ip=access_ip,
                record_count=record_count,
                file_size=file_size,
                status=status,
                error_message=error_message,
                response_time_ms=response_time_ms,
                access_time=datetime.now()
            )
            db.add(log)
            await db.commit()
        except Exception as e:
            logger.error(f"记录访问日志失败: {e}")

    def _get_file_extension(self, format: str) -> str:
        """获取文件扩展名"""
        extensions = {
            'excel': '.xlsx',
            'csv': '.csv',
            'json': '.json'
        }
        return extensions.get(format, '.txt')

    async def _generate_excel_file(self, file_path: Path, data: List[Dict]) -> int:
        """生成Excel文件"""
        import openpyxl

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            if data:
                # 写入表头
                headers = list(data[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)

                # 写入数据
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

            wb.save(str(file_path))

            return file_path.stat().st_size

        except Exception as e:
            logger.error(f"生成Excel文件失败: {e}")
            raise

    async def _generate_csv_file(self, file_path: Path, data: List[Dict]) -> int:
        """生成CSV文件"""
        import csv

        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                if data:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)

            return file_path.stat().st_size

        except Exception as e:
            logger.error(f"生成CSV文件失败: {e}")
            raise

    async def _generate_json_file(self, file_path: Path, data: List[Dict]) -> int:
        """生成JSON文件"""
        import json

        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)

            return file_path.stat().st_size

        except Exception as e:
            logger.error(f"生成JSON文件失败: {e}")
            raise

    def _calculate_popularity_score(
            self,
            preview_count: int,
            download_count: int,
            api_call_count: int
    ) -> float:
        """
        计算热度评分

        简单算法：预览×1 + 下载×5 + API调用×3
        归一化到0-100
        """
        try:
            raw_score = (
                    preview_count * 1 +
                    download_count * 5 +
                    api_call_count * 3
            )

            # 简单的对数归一化
            import math
            if raw_score == 0:
                return 0.0

            normalized = min(100.0, (math.log10(raw_score + 1) / math.log10(1000)) * 100)
            return round(normalized, 2)

        except Exception:
            return 0.0

    async def import_tables_from_datasource(
            self,
            db: AsyncSession,
            data_source_id: int,
            catalog_id: int,
            database_name: Optional[str] = None,
            table_patterns: Optional[List[str]] = None,
            include_columns: bool = True,
            creator: Optional[str] = None
    ) -> Tuple[List[DataAsset], List[Dict[str, Any]]]:
        """
        从数据源批量导入表作为资产

        Args:
            db: 数据库会话
            data_source_id: 数据源ID
            catalog_id: 目标目录ID
            database_name: 数据库名
            table_patterns: 表名匹配模式列表，如: ['cus_%', 'order_%']
            include_columns: 是否导入字段信息
            creator: 创建人

        Returns:
            (成功创建的资产列表, 失败的项目列表)
        """
        success_assets = []
        failed_items = []

        try:
            # 1. 验证数据源
            data_source = await self._get_data_source(db, data_source_id)
            if not data_source:
                raise ValueError(f"数据源 ID {data_source_id} 不存在")

            # 2. 验证目录
            catalog = await data_catalog_service.get_catalog_by_id(db, catalog_id)
            if not catalog:
                raise ValueError(f"目录 ID {catalog_id} 不存在")

            # 3. 获取数据源的表列表
            from app.services.optimized_data_integration_service import get_optimized_data_integration_service
            integration_service = get_optimized_data_integration_service()

            logger.info(f"开始从数据源 {data_source.name} 获取表列表")

            tables_result = await integration_service.get_tables(
                data_source.name,
                database=database_name,
                limit=1000  # 限制最多1000张表
            )

            if not tables_result.get('success'):
                raise ValueError(f"获取表列表失败: {tables_result.get('error')}")

            tables = tables_result.get('tables', [])
            logger.info(f"获取到 {len(tables)} 张表")

            # 4. 过滤表（根据table_patterns）
            filtered_tables = []
            if table_patterns:
                for table in tables:
                    table_name = table.get('table_name', '')
                    # 检查表名是否匹配任一模式
                    for pattern in table_patterns:
                        # 将SQL LIKE模式转换为Python正则
                        import re
                        regex_pattern = pattern.replace('%', '.*').replace('_', '.')
                        if re.match(f"^{regex_pattern}$", table_name):
                            filtered_tables.append(table)
                            break
            else:
                filtered_tables = tables

            logger.info(f"过滤后剩余 {len(filtered_tables)} 张表")

            # 5. 批量创建资产
            for idx, table_info in enumerate(filtered_tables):
                try:
                    table_name = table_info.get('table_name', '')

                    # 生成资产编码（使用表名）
                    asset_code = f"{data_source.name.upper()}_{table_name.upper()}"

                    # 检查是否已存在
                    existing = await self._get_by_code(db, asset_code)
                    if existing:
                        failed_items.append({
                            "index": idx + 1,
                            "table_name": table_name,
                            "error": f"资产编码 {asset_code} 已存在"
                        })
                        continue

                    # 创建资产
                    asset = DataAsset(
                        asset_name=table_info.get('comment') or table_name,
                        asset_code=asset_code,
                        asset_type="table",
                        catalog_id=catalog_id,
                        data_source_id=data_source_id,
                        database_name=database_name or table_info.get('database'),
                        table_name=table_name,
                        business_description=table_info.get('comment'),
                        quality_level="C",
                        status="normal",
                        is_public=True,
                        is_api_published=False,
                        preview_count=0,
                        download_count=0,
                        api_call_count=0,
                        allow_download=True,
                        max_download_rows=10000,
                        created_by=creator,
                        updated_by=creator
                    )

                    db.add(asset)
                    await db.flush()  # 获取ID

                    # 6. 如果需要导入字段信息
                    if include_columns:
                        try:
                            await self._fetch_and_save_metadata(db, asset, data_source)
                        except Exception as e:
                            logger.warning(f"获取表 {table_name} 元数据失败: {e}")

                    success_assets.append(asset)

                    # 每处理10个提交一次
                    if len(success_assets) % 10 == 0:
                        await db.commit()
                        logger.info(f"已处理 {len(success_assets)} 张表")

                except Exception as e:
                    logger.error(f"导入表 {table_name} 失败: {e}")
                    failed_items.append({
                        "index": idx + 1,
                        "table_name": table_name,
                        "error": str(e)
                    })

            # 7. 最后提交
            await db.commit()

            # 8. 更新目录的资产计数
            if success_assets:
                await data_catalog_service.update_asset_count(
                    db,
                    catalog_id,
                    increment=len(success_assets)
                )

            logger.info(
                f"批量导入完成: 成功 {len(success_assets)}，失败 {len(failed_items)}"
            )

            return success_assets, failed_items

        except ValueError:
            raise
        except Exception as e:
            await db.rollback()
            logger.error(f"批量导入表失败: {e}")
            raise ValueError(f"批量导入失败: {str(e)}")

    # ==================== 私有方法 ====================

    async def _get_by_code(
            self,
            db: AsyncSession,
            asset_code: str
    ) -> Optional[DataAsset]:
        """根据编码获取资产"""
        try:
            result = await db.execute(
                select(DataAsset).where(DataAsset.asset_code == asset_code)
            )
            return result.scalar_one_or_none()
        except Exception:
            return None

    async def _get_data_source(
            self,
            db: AsyncSession,
            data_source_id: int
    ) -> Optional[DataSource]:
        """获取数据源"""
        try:
            result = await db.execute(
                select(DataSource).where(DataSource.id == data_source_id)
            )
            return result.scalar_one_or_none()
        except Exception:
            return None

    async def _fetch_and_save_metadata(
            self,
            db: AsyncSession,
            asset: DataAsset,
            data_source: DataSource
    ) -> None:
        """
        从数据源获取表元数据并保存

        Args:
            db: 数据库会话
            asset: 资产对象
            data_source: 数据源对象
        """
        try:
            # TODO: 这里需要调用数据集成服务获取表元数据
            # 暂时先模拟获取
            logger.info(f"开始获取表元数据: {asset.table_name}")

            # 调用数据集成服务获取表结构
            from app.services.optimized_data_integration_service import get_optimized_data_integration_service
            integration_service = get_optimized_data_integration_service()

            # 获取表结构
            schema_result = await integration_service.get_table_schema(
                data_source.name,
                asset.table_name,
                asset.database_name
            )

            if not schema_result.get('success'):
                logger.warning(f"获取表结构失败: {schema_result.get('error')}")
                return

            schema_data = schema_result.get('schema', {})
            columns_data = schema_data.get('columns', [])

            # 保存字段信息
            for idx, col_data in enumerate(columns_data):
                column = AssetColumn(
                    asset_id=asset.id,
                    column_name=col_data.get('name', ''),
                    data_type=col_data.get('type', ''),
                    length=col_data.get('max_length'),
                    precision=col_data.get('precision'),
                    scale=col_data.get('scale'),
                    is_primary_key=col_data.get('key') == 'PRI',
                    is_nullable=col_data.get('nullable', True),
                    default_value=col_data.get('default'),
                    business_description=col_data.get('comment'),
                    sort_order=idx
                )
                db.add(column)

            # 更新资产的字段数量
            asset.column_count = len(columns_data)

            # 获取统计信息
            stats_result = await integration_service.get_table_statistics(
                data_source.name,
                asset.table_name,
                asset.database_name
            )

            if stats_result.get('success'):
                stats = stats_result.get('statistics', {})
                asset.row_count = stats.get('row_count')
                asset.data_size = stats.get('estimated_size')

            logger.info(f"表元数据获取成功: {len(columns_data)} 个字段")

        except Exception as e:
            logger.error(f"获取表元数据失败: {e}")
            # 不抛出异常，允许在没有元数据的情况下创建资产


# 创建全局服务实例
data_asset_service = DataAssetService()