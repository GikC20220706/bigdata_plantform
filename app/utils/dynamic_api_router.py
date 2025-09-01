"""
动态API路由管理器 - 支持运行时注册和注销API路由
"""

import asyncio
import json
import time
from datetime import datetime
from typing import Dict, List, Any, Optional, Callable
from fastapi import APIRouter, Request, HTTPException, Query, Body, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from loguru import logger
from io import StringIO, BytesIO
import csv
import openpyxl

from app.models.custom_api import CustomAPI
from app.services.custom_api_service import custom_api_service
from app.utils.custom_api_cache import custom_api_cache
from app.utils.custom_api_rate_limiter import rate_limiter
from app.utils.database import get_async_db
from app.utils.response import create_response


class DynamicAPIRouter:
    """动态API路由管理器"""

    def __init__(self):
        self.router = APIRouter(prefix="/api/custom", tags=["动态生成的API"])
        self.registered_apis: Dict[str, Dict] = {}  # api_name -> api_info
        self.route_handlers: Dict[str, Callable] = {}  # api_path -> handler

    async def register_api(self, api: CustomAPI) -> bool:
        """注册单个API到动态路由"""
        try:
            # 检查是否已注册
            if api.api_name in self.registered_apis:
                logger.warning(f"API {api.api_name} 已存在，将覆盖注册")
                await self.unregister_api(api.api_name)

            # 创建动态端点处理函数
            endpoint_handler = self._create_endpoint_handler(api)

            # 提取路径（去掉 /api/custom 前缀）
            route_path = api.api_path.replace("/api/custom", "")
            if not route_path.startswith("/"):
                route_path = "/" + route_path

            # 注册路由
            self.router.add_api_route(
                path=route_path,
                endpoint=endpoint_handler,
                methods=[api.http_method.value],
                summary=f"{api.api_name}",
                description=api.description or f"动态生成的API: {api.api_name}",
                tags=[f"自定义API", api.data_source.name if api.data_source else "Unknown"],
                response_model=None,
                name=f"custom_api_{api.api_name}"
            )

            # 记录注册信息
            self.registered_apis[api.api_name] = {
                "api_id": api.id,
                "api_path": api.api_path,
                "route_path": route_path,
                "http_method": api.http_method.value,
                "registered_at": datetime.now(),
                "data_source_name": api.data_source.name if api.data_source else None
            }

            self.route_handlers[api.api_path] = endpoint_handler

            logger.info(f"成功注册动态API: {api.api_name} -> {api.api_path} ({api.http_method.value})")
            return True

        except Exception as e:
            logger.error(f"注册动态API失败 {api.api_name}: {e}")
            return False

    async def unregister_api(self, api_name: str) -> bool:
        """注销API路由"""
        try:
            if api_name not in self.registered_apis:
                logger.warning(f"API {api_name} 未注册，无需注销")
                return False

            api_info = self.registered_apis[api_name]
            api_path = api_info["api_path"]

            # 从路由处理器中移除
            if api_path in self.route_handlers:
                del self.route_handlers[api_path]

            # 从注册列表中移除
            del self.registered_apis[api_name]

            # 清理缓存
            await custom_api_cache.clear_api_cache(api_name)

            logger.info(f"成功注销动态API: {api_name}")
            return True

        except Exception as e:
            logger.error(f"注销动态API失败 {api_name}: {e}")
            return False

    async def refresh_all_apis(self, db: AsyncSession) -> Dict[str, Any]:
        """刷新所有动态API路由"""
        try:
            start_time = time.time()

            # 获取所有活跃的API
            apis, _ = await custom_api_service.get_apis_list(
                db=db,
                is_active=True,
                limit=1000  # 假设不会超过1000个API
            )

            # 统计信息
            registered_count = 0
            failed_count = 0
            updated_count = 0

            for api in apis:
                if api.api_name in self.registered_apis:
                    # 检查是否需要更新
                    existing_info = self.registered_apis[api.api_name]
                    if existing_info["api_id"] != api.id:
                        await self.unregister_api(api.api_name)
                        success = await self.register_api(api)
                        if success:
                            updated_count += 1
                        else:
                            failed_count += 1
                else:
                    # 新注册
                    success = await self.register_api(api)
                    if success:
                        registered_count += 1
                    else:
                        failed_count += 1

            # 清理已删除的API
            current_api_names = {api.api_name for api in apis}
            registered_names = set(self.registered_apis.keys())

            removed_count = 0
            for api_name in registered_names - current_api_names:
                if await self.unregister_api(api_name):
                    removed_count += 1

            elapsed_time = time.time() - start_time

            result = {
                "refresh_time": datetime.now(),
                "elapsed_seconds": round(elapsed_time, 3),
                "statistics": {
                    "total_apis": len(apis),
                    "registered": registered_count,
                    "updated": updated_count,
                    "removed": removed_count,
                    "failed": failed_count,
                    "active_routes": len(self.registered_apis)
                },
                "registered_apis": list(self.registered_apis.keys())
            }

            logger.info(f"动态路由刷新完成: {result['statistics']}")
            return result

        except Exception as e:
            logger.error(f"刷新动态API路由失败: {e}")
            raise

    def _create_endpoint_handler(self, api: CustomAPI) -> Callable:
        """创建动态端点处理函数"""

        async def dynamic_endpoint_handler(
                request: Request,
                db: AsyncSession = Depends(get_async_db)
        ):
            """动态生成的API端点处理函数"""
            start_time = time.time()
            client_ip = request.client.host if request.client else "unknown"
            user_agent = request.headers.get("user-agent", "unknown")

            try:
                # 1. 频率限制检查
                if not await rate_limiter.check_rate_limit(
                        api_id=api.id,
                        client_ip=client_ip,
                        limit=api.rate_limit
                ):
                    raise HTTPException(
                        status_code=429,
                        detail=f"请求频率超限，最大允许 {api.rate_limit} 次/分钟"
                    )

                # 2. 获取请求参数
                if api.http_method.value == "GET":
                    request_params = dict(request.query_params)
                elif api.http_method.value == "POST":
                    try:
                        request_params = await request.json()
                    except Exception:
                        request_params = {}
                else:
                    request_params = {}

                # 3. 检查缓存
                cache_key = f"api_{api.id}_{hash(str(sorted(request_params.items())))}"
                cached_result = await custom_api_cache.get_cached_result(
                    cache_key,
                    api.cache_ttl
                )

                if cached_result:
                    logger.debug(f"缓存命中: {api.api_name}")
                    response_data = cached_result
                    response_data["cached"] = True
                    response_data["response_time_ms"] = int((time.time() - start_time) * 1000)
                else:
                    # 4. 执行查询
                    result = await custom_api_service.execute_api_query(
                        db=db,
                        api_id=api.id,
                        request_params=request_params,
                        client_ip=client_ip,
                        user_agent=user_agent
                    )

                    if not result.success:
                        raise HTTPException(status_code=500, detail=result.error_message)

                    response_data = {
                        "success": True,
                        "data": result.data,
                        "total_count": result.total_count,
                        "response_time_ms": result.response_time_ms,
                        "cached": False,
                        "api_info": {
                            "api_name": api.api_name,
                            "api_version": "1.0",
                            "data_source": api.data_source.name if api.data_source else None
                        }
                    }

                    # 5. 缓存结果
                    if api.cache_ttl > 0:
                        await custom_api_cache.cache_result(
                            cache_key,
                            response_data,
                            api.cache_ttl
                        )

                # 6. 根据响应格式返回数据
                if api.response_format.value == "json":
                    return JSONResponse(content=response_data)
                elif api.response_format.value == "csv":
                    return self._generate_csv_response(response_data["data"], api.api_name)
                elif api.response_format.value == "excel":
                    return self._generate_excel_response(response_data["data"], api.api_name)
                else:
                    return JSONResponse(content=response_data)

            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"动态API执行失败 {api.api_name}: {e}")
                raise HTTPException(status_code=500, detail=str(e))

        # 设置函数名称，便于调试
        dynamic_endpoint_handler.__name__ = f"api_{api.api_name}_handler"
        return dynamic_endpoint_handler

    def _generate_csv_response(self, data: List[Dict], api_name: str) -> StreamingResponse:
        """生成CSV响应"""
        output = StringIO()

        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

        csv_content = output.getvalue()
        output.close()

        return StreamingResponse(
            io=StringIO(csv_content),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={api_name}.csv"}
        )

    def _generate_excel_response(self, data: List[Dict], api_name: str) -> StreamingResponse:
        """生成Excel响应"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = api_name

        if data:
            # 写入表头
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入数据
            for row, record in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=record.get(header))

        # 保存到字节流
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return StreamingResponse(
            io=excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={api_name}.xlsx"}
        )

    def get_registered_apis_info(self) -> Dict[str, Any]:
        """获取已注册的API信息"""
        return {
            "total_registered": len(self.registered_apis),
            "apis": self.registered_apis,
            "router_info": {
                "prefix": self.router.prefix,
                "tags": self.router.tags,
                "route_count": len(self.router.routes)
            }
        }


# 创建全局动态路由管理器实例
dynamic_api_router = DynamicAPIRouter()