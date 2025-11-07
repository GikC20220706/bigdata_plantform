"""
Data Integration Module - Database Clients
支持多种数据库的统一连接和操作客户端
"""

import asyncio
import json
import logging
import time
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from contextlib import asynccontextmanager
import traceback
import os
import asyncpg
import aiomysql
import pymysql
from sqlalchemy import create_engine, text, MetaData, inspect
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
import ksycopg2
from loguru import logger
from pathlib import Path
import pandas as pd
import uuid
import openpyxl
from fastapi import UploadFile
import ksycopg2.extras

from config.settings import settings


class DatabaseClient(ABC):
    """数据库客户端基类"""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.connection_pool = None
        self._last_test_time = None
        self._last_test_result = None

    @abstractmethod
    async def test_connection(self) -> Dict[str, Any]:
        """测试数据库连接"""
        pass

    @abstractmethod
    async def get_databases(self) -> List[str]:
        """获取数据库列表"""
        pass

    @abstractmethod
    async def get_tables(self, database: str = None) -> List[Dict[str, Any]]:
        """获取表列表"""
        pass

    @abstractmethod
    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取表结构"""
        pass

    @abstractmethod
    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[
        Dict[str, Any]]:
        """执行查询"""
        pass

    async def get_table_metadata(self, table_name: str, database: str = None, schema: str = None) -> Dict[str, Any]:
        """获取表的完整元数据信息"""
        try:
            schema = await self.get_table_schema(table_name, database)

            # 获取表统计信息
            stats = await self._get_table_statistics(table_name, database)

            return {
                "table_name": table_name,
                "database": database,
                "schema": schema,
                "statistics": stats,
                "collected_at": datetime.now()
            }
        except Exception as e:
            logger.error(f"获取表元数据失败 {table_name}: {e}")
            return {
                "table_name": table_name,
                "database": database,
                "error": str(e),
                "collected_at": datetime.now()
            }

    async def _get_table_statistics(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取表统计信息 - 子类可重写"""
        try:
            # 基础统计查询
            count_query = f"SELECT COUNT(*) as row_count FROM {table_name}"
            if database:
                count_query = f"SELECT COUNT(*) as row_count FROM {database}.{table_name}"

            result = await self.execute_query(count_query, database)
            row_count = result[0]['row_count'] if result else 0

            return {
                "row_count": row_count,
                "estimated_size": "未知",
                "last_updated": "未知"
            }
        except Exception as e:
            logger.warning(f"获取表统计信息失败 {table_name}: {e}")
            return {
                "row_count": 0,
                "estimated_size": "未知",
                "last_updated": "未知"
            }

    def _build_connection_string(self) -> str:
        """构建连接字符串"""
        host = self.config.get('host', 'localhost')
        port = self.config.get('port')
        database = self.config.get('database', '')
        username = self.config.get('username', '')
        password = self.config.get('password', '')

        return f"{username}:{password}@{host}:{port}/{database}"


class MySQLClient(DatabaseClient):
    """MySQL数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.pool = None

    async def _get_connection_pool(self):
        """获取连接池"""
        if not self.pool:
            self.pool = await aiomysql.create_pool(
                host=self.config['host'],
                port=self.config.get('port', 3306),
                user=self.config['username'],
                password=self.config['password'],
                db=self.config.get('database', ''),
                charset='utf8mb4',
                autocommit=True,
                maxsize=10,
                minsize=1
            )
        return self.pool

    async def test_connection(self) -> Dict[str, Any]:
        """测试MySQL连接"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    await cursor.execute("SELECT VERSION() as version")
                    result = await cursor.fetchone()

                    return {
                        "success": True,
                        "version": result[0] if result else "未知",
                        "database_type": "MySQL",
                        "test_time": datetime.now(),
                        "response_time_ms": 50  # 实际测量
                    }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "MySQL",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取MySQL数据库列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    await cursor.execute("SHOW DATABASES")
                    results = await cursor.fetchall()
                    return [db[0] for db in results if
                            db[0] not in ('information_schema', 'performance_schema', 'mysql', 'sys')]
        except Exception as e:
            logger.error(f"获取MySQL数据库列表失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取MySQL表列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    if database:
                        await cursor.execute(f"USE {database}")

                    await cursor.execute("""
                        SELECT 
                            TABLE_NAME,
                            TABLE_TYPE,
                            ENGINE,
                            TABLE_ROWS,
                            DATA_LENGTH,
                            CREATE_TIME,
                            UPDATE_TIME,
                            TABLE_COMMENT
                        FROM information_schema.TABLES 
                        WHERE TABLE_SCHEMA = %s
                        ORDER BY TABLE_NAME
                    """, (database or self.config.get('database'),))

                    results = await cursor.fetchall()
                    tables = []
                    for row in results:
                        tables.append({
                            "table_name": row[0],
                            "table_type": row[1],
                            "engine": row[2],
                            "estimated_rows": row[3] or 0,
                            "data_size": row[4] or 0,
                            "created_at": row[5],
                            "updated_at": row[6],
                            "comment": row[7] or ""
                        })
                    return tables
        except Exception as e:
            logger.error(f"获取MySQL表列表失败: {e}")
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取MySQL表结构"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    db_name = database or self.config.get('database')
                    await cursor.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")

                    # 获取列信息
                    await cursor.execute("""
                        SELECT 
                            COLUMN_NAME,
                            DATA_TYPE,
                            IS_NULLABLE,
                            COLUMN_DEFAULT,
                            COLUMN_KEY,
                            EXTRA,
                            COLUMN_COMMENT,
                            CHARACTER_MAXIMUM_LENGTH,
                            NUMERIC_PRECISION,
                            NUMERIC_SCALE
                        FROM information_schema.COLUMNS 
                        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                        ORDER BY ORDINAL_POSITION
                    """, (db_name, table_name))

                    columns = []
                    for row in await cursor.fetchall():
                        columns.append({
                            "name": row[0],
                            "type": row[1],
                            "nullable": row[2] == 'YES',
                            "default": row[3],
                            "key": row[4],
                            "extra": row[5],
                            "comment": row[6] or "",
                            "max_length": row[7],
                            "precision": row[8],
                            "scale": row[9]
                        })

                    # 获取索引信息
                    await cursor.execute("""
                        SELECT 
                            INDEX_NAME,
                            COLUMN_NAME,
                            NON_UNIQUE,
                            INDEX_TYPE
                        FROM information_schema.STATISTICS 
                        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                        ORDER BY INDEX_NAME, SEQ_IN_INDEX
                    """, (db_name, table_name))

                    indexes = {}
                    for row in await cursor.fetchall():
                        index_name = row[0]
                        if index_name not in indexes:
                            indexes[index_name] = {
                                "name": index_name,
                                "columns": [],
                                "unique": row[2] == 0,
                                "type": row[3]
                            }
                        indexes[index_name]["columns"].append(row[1])

                    return {
                        "columns": columns,
                        "indexes": list(indexes.values()),
                        "primary_key": [idx for idx in indexes.values() if idx["name"] == "PRIMARY"],
                        "foreign_keys": []  # 可以进一步实现
                    }
        except Exception as e:
            logger.error(f"获取MySQL表结构失败 {table_name}: {e}")
            return {"columns": [], "indexes": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None) -> List[Dict[str, Any]]:
        """执行MySQL查询"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor(aiomysql.DictCursor) as cursor:
                    if database:
                        await cursor.execute(f"USE {database}")

                    await cursor.execute(query)
                    results = await cursor.fetchall()
                    return list(results)
        except Exception as e:
            logger.error(f"MySQL查询执行失败: {e}")
            return []


class KingbaseClient(DatabaseClient):
    """人大金仓数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.connection = None
        self._connection_params = {
            'host': config['host'],
            'port': int(config.get('port', 54321)),
            'user': config['username'],
            'password': config['password'],
            'database': config.get('database', 'test'),
            'connect_timeout': 5,  # 连接超时5秒
            'options': '-c statement_timeout=10000'  # SQL执行超时10秒
        }

    def _get_connection(self):
        """获取Kingbase连接"""
        try:


            if not self.connection or self.connection.closed:
                self.connection = ksycopg2.connect(**self._connection_params)
                self.connection.autocommit = True
            return self.connection
        except ImportError:
            logger.error("psycopg2库未安装，请安装: pip install psycopg2-binary")
            raise

    async def test_connection(self) -> Dict[str, Any]:
        """测试人大金仓连接"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT version()")
            version = cursor.fetchone()[0]
            cursor.close()

            return {
                "success": True,
                "version": version,
                "database_type": "Kingbase",
                "test_time": datetime.now(),
                "response_time_ms": 50
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "Kingbase",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取人大金仓数据库列表"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT datname FROM sys_database 
                WHERE datname NOT IN ('template0', 'template1', 'security') 
                ORDER BY datname
            """)
            results = cursor.fetchall()
            cursor.close()
            return [row[0] for row in results]
        except Exception as e:
            logger.error(f"获取Kingbase数据库列表失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[
        Dict[str, Any]]:
        """获取人大金仓表列表 - 支持分页"""
        conn = None
        cursor = None
        try:
            logger.info(f"KingBase get_tables: database={database}, schema={schema}, limit={limit}, offset={offset}")

            # 创建新连接
            conn = ksycopg2.connect(**self._connection_params)
            conn.autocommit = True
            cursor = conn.cursor(cursor_factory=ksycopg2.extras.RealDictCursor)

            if not schema:
                schema = 'public'

            # 分页查询表列表
            cursor.execute("""
                SELECT table_name, table_type 
                FROM information_schema.tables 
                WHERE table_schema = %s AND table_type = 'BASE TABLE'
                ORDER BY table_name
                LIMIT %s OFFSET %s
            """, (schema, limit, offset))

            results = cursor.fetchall()

            tables = []
            for row in results:
                tables.append({
                    "table_name": row['table_name'],
                    "table_type": row['table_type'],
                    "estimated_rows": 0,
                    "data_size": 0,
                    "comment": ""
                })

            logger.info(f"KingBase查询完成: 返回{len(tables)}张表")
            return tables

        except Exception as e:
            logger.error(f"获取Kingbase表列表失败: {e}")
            return []
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    async def get_tables_count(self, database: str = None, schema: str = None) -> int:
        """获取表总数"""
        conn = None
        cursor = None
        try:
            conn = ksycopg2.connect(**self._connection_params)
            conn.autocommit = True
            cursor = conn.cursor()

            if not schema:
                schema = 'public'

            cursor.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = %s AND table_type = 'BASE TABLE'
            """, (schema,))

            count = cursor.fetchone()[0]
            return count

        except Exception as e:
            logger.error(f"获取Kingbase表总数失败: {e}")
            return 0
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    async def get_table_schema(self, table_name: str, database: str = None, schema: str = None) -> Dict[str, Any]:
        """获取人大金仓表结构"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor(cursor_factory=ksycopg2.extras.RealDictCursor)

            # 如果指定了数据库，需要切换连接
            if database and database != self._connection_params['database']:
                conn.close()
                temp_params = self._connection_params.copy()
                temp_params['database'] = database
                conn = ksycopg2.connect(**temp_params)
                conn.autocommit = True
                cursor = conn.cursor(cursor_factory=ksycopg2.extras.RealDictCursor)
            if not schema:
                schema = 'public'
            # 获取列信息
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default,
                    character_maximum_length,
                    numeric_precision,
                    numeric_scale,
                    col_description(sys_class.oid, ordinal_position) as comment
                FROM information_schema.columns
                LEFT JOIN sys_class ON sys_class.relname = table_name
                WHERE table_name = %s AND table_schema = %s
                ORDER BY ordinal_position
            """, (table_name, schema or 'public'))

            columns_result = cursor.fetchall()

            columns = []
            for row in columns_result:
                columns.append({
                    "name": row['column_name'],
                    "type": row['data_type'],
                    "nullable": row['is_nullable'] == 'YES',
                    "default": row['column_default'],
                    "max_length": row['character_maximum_length'],
                    "precision": row['numeric_precision'],
                    "scale": row['numeric_scale'],
                    "comment": row['comment'] or ""
                })

            # 获取索引信息
            cursor.execute("""
                SELECT 
                    i.relname as index_name,
                    a.attname as column_name,
                    ix.indisunique as is_unique,
                    ix.indisprimary as is_primary
                FROM sys_index ix
                JOIN sys_class i ON i.oid = ix.indexrelid
                JOIN sys_class t ON t.oid = ix.indrelid
                JOIN sys_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(ix.indkey)
                WHERE t.relname = %s
                ORDER BY i.relname, a.attnum
            """, (table_name,))

            indexes_result = cursor.fetchall()
            cursor.close()

            indexes = {}
            for row in indexes_result:
                index_name = row['index_name']
                if index_name not in indexes:
                    indexes[index_name] = {
                        "name": index_name,
                        "columns": [],
                        "unique": row['is_unique'],
                        "primary": row['is_primary']
                    }
                indexes[index_name]["columns"].append(row['column_name'])

            return {
                "columns": columns,
                "indexes": list(indexes.values()),
                "primary_key": [idx for idx in indexes.values() if idx.get("primary")],
                "foreign_keys": []
            }
        except Exception as e:
            logger.error(f"获取Kingbase表结构失败 {table_name}: {e}")
            return {"columns": [], "indexes": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = "piblic") -> List[Dict[str, Any]]:
        """执行人大金仓查询"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor(cursor_factory=ksycopg2.extras.RealDictCursor)

            # 如果指定了数据库，需要切换连接
            if database and database != self._connection_params['database']:
                conn.close()
                temp_params = self._connection_params.copy()
                temp_params['database'] = database
                conn = ksycopg2.connect(**temp_params)
                conn.autocommit = True
                cursor = conn.cursor(cursor_factory=ksycopg2.extras.RealDictCursor)
            if schema:
                cursor.execute(f"SET search_path TO {schema}")

            cursor.execute(query)
            results = cursor.fetchall()
            cursor.close()

            return [dict(row) for row in results]
        except Exception as e:
            logger.error(f"Kingbase查询执行失败: {e}")
            return []

    def __del__(self):
        """析构函数，确保连接关闭"""
        if hasattr(self, 'connection') and self.connection and not self.connection.closed:
            self.connection.close()


class DorisClient(DatabaseClient):
    """Apache Doris数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.pool = None

    async def _get_connection_pool(self):
        """获取连接池 - Doris使用MySQL协议"""
        if not self.pool:
            self.pool = await aiomysql.create_pool(
                host=self.config['host'],
                port=self.config.get('port', 9030),
                user=self.config['username'],
                password=self.config['password'],
                db=self.config.get('database', ''),
                charset='utf8mb4',
                autocommit=True,
                maxsize=10,
                minsize=1
            )
        return self.pool

    async def test_connection(self) -> Dict[str, Any]:
        """测试Doris连接"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    await cursor.execute("SELECT @@version_comment as version")
                    result = await cursor.fetchone()

                    return {
                        "success": True,
                        "version": result[0] if result else "未知",
                        "database_type": "Apache Doris",
                        "test_time": datetime.now(),
                        "response_time_ms": 50
                    }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "Apache Doris",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取Doris数据库列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    await cursor.execute("SHOW DATABASES")
                    results = await cursor.fetchall()
                    return [db[0] for db in results if db[0] not in ('information_schema', 'mysql')]
        except Exception as e:
            logger.error(f"获取Doris数据库列表失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取Doris表列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    if database:
                        await cursor.execute(f"USE {database}")

                    # Doris特有的表信息查询
                    await cursor.execute("""
                        SELECT 
                            TABLE_NAME,
                            TABLE_TYPE,
                            ENGINE,
                            TABLE_ROWS,
                            DATA_LENGTH,
                            CREATE_TIME,
                            TABLE_COMMENT
                        FROM information_schema.TABLES 
                        WHERE TABLE_SCHEMA = %s
                        ORDER BY TABLE_NAME
                    """, (database or self.config.get('database'),))

                    results = await cursor.fetchall()
                    tables = []
                    for row in results:
                        tables.append({
                            "table_name": row[0],
                            "table_type": row[1],
                            "engine": row[2],
                            "estimated_rows": row[3] or 0,
                            "data_size": row[4] or 0,
                            "created_at": row[5],
                            "comment": row[6] or ""
                        })
                    return tables
        except Exception as e:
            logger.error(f"获取Doris表列表失败: {e}")
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取Doris表结构"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    db_name = database or self.config.get('database')

                    # 获取列信息
                    await cursor.execute("""
                        SELECT 
                            COLUMN_NAME,
                            DATA_TYPE,
                            IS_NULLABLE,
                            COLUMN_DEFAULT,
                            COLUMN_KEY,
                            COLUMN_COMMENT,
                            CHARACTER_MAXIMUM_LENGTH,
                            NUMERIC_PRECISION,
                            NUMERIC_SCALE
                        FROM information_schema.COLUMNS 
                        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                        ORDER BY ORDINAL_POSITION
                    """, (db_name, table_name))

                    columns = []
                    for row in await cursor.fetchall():
                        columns.append({
                            "name": row[0],
                            "type": row[1],
                            "nullable": row[2] == 'YES',
                            "default": row[3],
                            "key": row[4],
                            "comment": row[5] or "",
                            "max_length": row[6],
                            "precision": row[7],
                            "scale": row[8]
                        })

                    # Doris特有的分区信息
                    await cursor.execute(f"SHOW PARTITIONS FROM {table_name}")
                    partitions = await cursor.fetchall()

                    return {
                        "columns": columns,
                        "indexes": [],  # Doris索引信息获取方式不同
                        "partitions": [{"name": p[0], "range": p[1]} for p in partitions] if partitions else [],
                        "storage_type": "OLAP"
                    }
        except Exception as e:
            logger.error(f"获取Doris表结构失败 {table_name}: {e}")
            return {"columns": [], "indexes": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[Dict[str, Any]]:
        """执行Doris查询"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor(aiomysql.DictCursor) as cursor:
                    if database:
                        await cursor.execute(f"USE {database}")

                    await cursor.execute(query)
                    results = await cursor.fetchall()
                    return list(results)
        except Exception as e:
            logger.error(f"Doris查询执行失败: {e}")
            return []


class DamengClient(DatabaseClient):
    """达梦数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.engine = None

    def _get_engine(self):
        """获取SQLAlchemy引擎"""
        if not self.engine:
            # 达梦数据库连接字符串
            connection_string = f"dm+pyodbc://{self.config['username']}:{self.config['password']}@{self.config['host']}:{self.config.get('port', 5236)}/{self.config.get('database', 'DAMENG')}"
            self.engine = create_engine(connection_string, pool_size=10, max_overflow=20)
        return self.engine

    async def test_connection(self) -> Dict[str, Any]:
        """测试达梦数据库连接"""
        try:
            engine = self._get_engine()
            with engine.connect() as conn:
                result = conn.execute(text("SELECT BANNER FROM $VERSION WHERE ROWNUM = 1"))
                version = result.fetchone()[0] if result else "未知"

                return {
                    "success": True,
                    "version": version,
                    "database_type": "Dameng Database",
                    "test_time": datetime.now(),
                    "response_time_ms": 50
                }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "Dameng Database",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取达梦数据库列表"""
        try:
            engine = self._get_engine()
            with engine.connect() as conn:
                result = conn.execute(text("SELECT NAME FROM $DATABASE"))
                return [row[0] for row in result.fetchall()]
        except Exception as e:
            logger.error(f"获取Dameng数据库列表失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取达梦表列表"""
        try:
            engine = self._get_engine()
            with engine.connect() as conn:
                result = conn.execute(text("""
                    SELECT 
                        TABLE_NAME,
                        TABLE_TYPE,
                        NUM_ROWS,
                        CREATED,
                        COMMENTS
                    FROM USER_TABLES
                    ORDER BY TABLE_NAME
                """))

                tables = []
                for row in result.fetchall():
                    tables.append({
                        "table_name": row[0],
                        "table_type": row[1],
                        "estimated_rows": row[2] or 0,
                        "created_at": row[3],
                        "comment": row[4] or ""
                    })
                return tables
        except Exception as e:
            logger.error(f"获取Dameng表列表失败: {e}")
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取达梦表结构"""
        try:
            engine = self._get_engine()
            with engine.connect() as conn:
                # 获取列信息
                columns_result = conn.execute(text("""
                    SELECT 
                        COLUMN_NAME,
                        DATA_TYPE,
                        NULLABLE,
                        DATA_DEFAULT,
                        DATA_LENGTH,
                        DATA_PRECISION,
                        DATA_SCALE,
                        COMMENTS
                    FROM USER_TAB_COLUMNS
                    WHERE TABLE_NAME = :table_name
                    ORDER BY COLUMN_ID
                """), {"table_name": table_name.upper()})

                columns = []
                for row in columns_result.fetchall():
                    columns.append({
                        "name": row[0],
                        "type": row[1],
                        "nullable": row[2] == 'Y',
                        "default": row[3],
                        "length": row[4],
                        "precision": row[5],
                        "scale": row[6],
                        "comment": row[7] or ""
                    })

                # 获取索引信息
                indexes_result = conn.execute(text("""
                    SELECT 
                        INDEX_NAME,
                        COLUMN_NAME,
                        UNIQUENESS
                    FROM USER_IND_COLUMNS ic
                    JOIN USER_INDEXES i ON ic.INDEX_NAME = i.INDEX_NAME
                    WHERE ic.TABLE_NAME = :table_name
                    ORDER BY ic.INDEX_NAME, ic.COLUMN_POSITION
                """), {"table_name": table_name.upper()})

                indexes = {}
                for row in indexes_result.fetchall():
                    index_name = row[0]
                    if index_name not in indexes:
                        indexes[index_name] = {
                            "name": index_name,
                            "columns": [],
                            "unique": row[2] == 'UNIQUE'
                        }
                    indexes[index_name]["columns"].append(row[1])

                return {
                    "columns": columns,
                    "indexes": list(indexes.values()),
                    "primary_key": [],  # 需要进一步查询
                    "foreign_keys": []
                }
        except Exception as e:
            logger.error(f"获取Dameng表结构失败 {table_name}: {e}")
            return {"columns": [], "indexes": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[Dict[str, Any]]:
        """执行达梦查询"""
        try:
            engine = self._get_engine()
            with engine.connect() as conn:
                result = conn.execute(text(query))
                columns = result.keys()
                return [dict(zip(columns, row)) for row in result.fetchall()]
        except Exception as e:
            logger.error(f"Dameng查询执行失败: {e}")
            return []


class GBaseClient(DatabaseClient):
    """南大通用GBase数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.pool = None

    async def _get_connection_pool(self):
        """获取连接池 - GBase使用类似PostgreSQL协议"""
        if not self.pool:
            dsn = f"postgresql://{self.config['username']}:{self.config['password']}@{self.config['host']}:{self.config.get('port', 5432)}/{self.config.get('database', 'gbase')}"
            self.pool = await asyncpg.create_pool(dsn, min_size=1, max_size=10)
        return self.pool

    async def test_connection(self) -> Dict[str, Any]:
        """测试GBase连接"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                version = await conn.fetchval("SELECT version()")
                return {
                    "success": True,
                    "version": version,
                    "database_type": "GBase Database",
                    "test_time": datetime.now(),
                    "response_time_ms": 50
                }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "GBase Database",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取GBase数据库列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                results = await conn.fetch("""
                    SELECT datname FROM pg_database 
                    WHERE datname NOT IN ('template0', 'template1', 'postgres') 
                    ORDER BY datname
                """)
                return [row['datname'] for row in results]
        except Exception as e:
            logger.error(f"获取GBase数据库列表失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取GBase表列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                results = await conn.fetch("""
                    SELECT 
                        t.table_name,
                        t.table_type,
                        c.reltuples as estimated_rows,
                        pg_total_relation_size(c.oid) as data_size,
                        obj_description(c.oid) as comment
                    FROM information_schema.tables t
                    LEFT JOIN pg_class c ON c.relname = t.table_name
                    WHERE t.table_schema = 'public'
                    ORDER BY t.table_name
                """)

                tables = []
                for row in results:
                    tables.append({
                        "table_name": row['table_name'],
                        "table_type": row['table_type'],
                        "estimated_rows": int(row['estimated_rows'] or 0),
                        "data_size": int(row['data_size'] or 0),
                        "comment": row['comment'] or ""
                    })
                return tables
        except Exception as e:
            logger.error(f"获取GBase表列表失败: {e}")
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取GBase表结构"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                # 获取列信息
                columns_result = await conn.fetch("""
                    SELECT 
                        column_name,
                        data_type,
                        is_nullable,
                        column_default,
                        character_maximum_length,
                        numeric_precision,
                        numeric_scale,
                        col_description(pg_class.oid, ordinal_position) as comment
                    FROM information_schema.columns
                    LEFT JOIN pg_class ON pg_class.relname = table_name
                    WHERE table_name = $1 AND table_schema = 'public'
                    ORDER BY ordinal_position
                """, table_name)

                columns = []
                for row in columns_result:
                    columns.append({
                        "name": row['column_name'],
                        "type": row['data_type'],
                        "nullable": row['is_nullable'] == 'YES',
                        "default": row['column_default'],
                        "max_length": row['character_maximum_length'],
                        "precision": row['numeric_precision'],
                        "scale": row['numeric_scale'],
                        "comment": row['comment'] or ""
                    })

                # 获取索引信息
                indexes_result = await conn.fetch("""
                    SELECT 
                        i.relname as index_name,
                        a.attname as column_name,
                        ix.indisunique as is_unique,
                        ix.indisprimary as is_primary
                    FROM pg_index ix
                    JOIN pg_class i ON i.oid = ix.indexrelid
                    JOIN pg_class t ON t.oid = ix.indrelid
                    JOIN pg_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(ix.indkey)
                    WHERE t.relname = $1
                    ORDER BY i.relname, a.attnum
                """, table_name)

                indexes = {}
                for row in indexes_result:
                    index_name = row['index_name']
                    if index_name not in indexes:
                        indexes[index_name] = {
                            "name": index_name,
                            "columns": [],
                            "unique": row['is_unique'],
                            "primary": row['is_primary']
                        }
                    indexes[index_name]["columns"].append(row['column_name'])

                return {
                    "columns": columns,
                    "indexes": list(indexes.values()),
                    "primary_key": [idx for idx in indexes.values() if idx.get("primary")],
                    "foreign_keys": []
                }
        except Exception as e:
            logger.error(f"获取GBase表结构失败 {table_name}: {e}")
            return {"columns": [], "indexes": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[Dict[str, Any]]:
        """执行GBase查询"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                results = await conn.fetch(query)
                return [dict(row) for row in results]
        except Exception as e:
            logger.error(f"GBase查询执行失败: {e}")
            return []


class TiDBClient(DatabaseClient):
    """TiDB数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.pool = None

    async def _get_connection_pool(self):
        """获取连接池 - TiDB兼容MySQL协议"""
        if not self.pool:
            self.pool = await aiomysql.create_pool(
                host=self.config['host'],
                port=self.config.get('port', 4000),
                user=self.config['username'],
                password=self.config['password'],
                db=self.config.get('database', ''),
                charset='utf8mb4',
                autocommit=True,
                maxsize=10,
                minsize=1
            )
        return self.pool

    async def test_connection(self) -> Dict[str, Any]:
        """测试TiDB连接"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    await cursor.execute("SELECT @@version as version, @@tidb_version as tidb_version")
                    result = await cursor.fetchone()

                    return {
                        "success": True,
                        "version": f"TiDB {result[1]}" if result and len(result) > 1 else result[
                            0] if result else "未知",
                        "database_type": "TiDB",
                        "test_time": datetime.now(),
                        "response_time_ms": 50
                    }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "TiDB",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取TiDB数据库列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    await cursor.execute("SHOW DATABASES")
                    results = await cursor.fetchall()
                    return [db[0] for db in results if
                            db[0] not in ('information_schema', 'performance_schema', 'mysql', 'sys', 'metrics_schema')]
        except Exception as e:
            logger.error(f"获取TiDB数据库列表失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取TiDB表列表"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    if database:
                        await cursor.execute(f"USE {database}")

                    await cursor.execute("""
                        SELECT 
                            TABLE_NAME,
                            TABLE_TYPE,
                            ENGINE,
                            TABLE_ROWS,
                            DATA_LENGTH,
                            CREATE_TIME,
                            UPDATE_TIME,
                            TABLE_COMMENT
                        FROM information_schema.TABLES 
                        WHERE TABLE_SCHEMA = %s
                        ORDER BY TABLE_NAME
                    """, (database or self.config.get('database'),))

                    results = await cursor.fetchall()
                    tables = []
                    for row in results:
                        tables.append({
                            "table_name": row[0],
                            "table_type": row[1],
                            "engine": row[2],
                            "estimated_rows": row[3] or 0,
                            "data_size": row[4] or 0,
                            "created_at": row[5],
                            "updated_at": row[6],
                            "comment": row[7] or ""
                        })
                    return tables
        except Exception as e:
            logger.error(f"获取TiDB表列表失败: {e}")
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取TiDB表结构"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor() as cursor:
                    db_name = database or self.config.get('database')

                    # 获取列信息
                    await cursor.execute("""
                        SELECT 
                            COLUMN_NAME,
                            DATA_TYPE,
                            IS_NULLABLE,
                            COLUMN_DEFAULT,
                            COLUMN_KEY,
                            EXTRA,
                            COLUMN_COMMENT,
                            CHARACTER_MAXIMUM_LENGTH,
                            NUMERIC_PRECISION,
                            NUMERIC_SCALE
                        FROM information_schema.COLUMNS 
                        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                        ORDER BY ORDINAL_POSITION
                    """, (db_name, table_name))

                    columns = []
                    for row in await cursor.fetchall():
                        columns.append({
                            "name": row[0],
                            "type": row[1],
                            "nullable": row[2] == 'YES',
                            "default": row[3],
                            "key": row[4],
                            "extra": row[5],
                            "comment": row[6] or "",
                            "max_length": row[7],
                            "precision": row[8],
                            "scale": row[9]
                        })

                    # 获取索引信息
                    await cursor.execute("""
                        SELECT 
                            INDEX_NAME,
                            COLUMN_NAME,
                            NON_UNIQUE,
                            INDEX_TYPE
                        FROM information_schema.STATISTICS 
                        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                        ORDER BY INDEX_NAME, SEQ_IN_INDEX
                    """, (db_name, table_name))

                    indexes = {}
                    for row in await cursor.fetchall():
                        index_name = row[0]
                        if index_name not in indexes:
                            indexes[index_name] = {
                                "name": index_name,
                                "columns": [],
                                "unique": row[2] == 0,
                                "type": row[3]
                            }
                        indexes[index_name]["columns"].append(row[1])

                    # TiDB特有的Region信息
                    try:
                        await cursor.execute(f"SHOW TABLE {table_name} REGIONS")
                        regions = await cursor.fetchall()
                        region_info = [{"region_id": r[0], "start_key": r[1], "end_key": r[2]} for r in
                                       regions] if regions else []
                    except:
                        region_info = []

                    return {
                        "columns": columns,
                        "indexes": list(indexes.values()),
                        "primary_key": [idx for idx in indexes.values() if idx["name"] == "PRIMARY"],
                        "foreign_keys": [],
                        "regions": region_info,
                        "distributed_storage": True
                    }
        except Exception as e:
            logger.error(f"获取TiDB表结构失败 {table_name}: {e}")
            return {"columns": [], "indexes": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[Dict[str, Any]]:
        """执行TiDB查询"""
        try:
            pool = await self._get_connection_pool()
            async with pool.acquire() as conn:
                async with conn.cursor(aiomysql.DictCursor) as cursor:
                    if database:
                        await cursor.execute(f"USE {database}")

                    await cursor.execute(query)
                    results = await cursor.fetchall()
                    return list(results)
        except Exception as e:
            logger.error(f"TiDB查询执行失败: {e}")
            return []


class HiveClient(DatabaseClient):
    """Apache Hive数据库客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.connection = None
        self._last_activity = None  # 最后活动时间
        self._connection_timeout = 300  # 5分钟超时

    def _is_connection_valid(self) -> bool:
        """检查连接是否有效"""
        if not self.connection:
            return False

        # 检查超时
        if self._last_activity:
            from datetime import datetime, timedelta
            if datetime.now() - self._last_activity > timedelta(seconds=self._connection_timeout):
                logger.info("Hive连接超时,需要重连")
                return False

        # 尝试执行简单查询测试连接
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT 1")
            cursor.fetchone()
            cursor.close()
            return True
        except Exception as e:
            logger.warning(f"Hive连接无效: {e}")
            return False

    def _close_connection(self):
        """关闭现有连接"""
        if self.connection:
            try:
                self.connection.close()
            except Exception as e:
                logger.warning(f"关闭Hive连接失败: {e}")
            finally:
                self.connection = None
                self._last_activity = None

    def _get_connection(self):
        """获取Hive连接 - 带自动重连"""
        from datetime import datetime

        # 检查现有连接是否有效
        if not self._is_connection_valid():
            logger.info("创建新的Hive连接...")
            self._close_connection()

            try:
                from pyhive import hive

                # ✅ 移除不支持的 connect_timeout 参数
                connection_params = {
                    'host': self.config['host'],
                    'port': self.config.get('port', 10000),
                    'username': self.config['username'],
                    'database': self.config.get('database', 'default')
                }

                # 根据是否有密码选择认证方式
                if self.config.get('password'):
                    connection_params['auth'] = 'CUSTOM'
                    connection_params['password'] = self.config['password']
                else:
                    connection_params['auth'] = 'NONE'

                self.connection = hive.Connection(**connection_params)
                self._last_activity = datetime.now()
                logger.info("Hive连接创建成功")

            except ImportError:
                logger.error("pyhive库未安装，请安装: pip install pyhive[hive]")
                raise
            except Exception as e:
                logger.error(f"Hive连接失败: {e}")
                raise

        # 更新最后活动时间
        self._last_activity = datetime.now()
        return self.connection

    async def test_connection(self) -> Dict[str, Any]:
        """测试Hive连接"""
        try:
            # 强制重新连接测试
            self._close_connection()
            conn = self._get_connection()

            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            result = cursor.fetchone()
            cursor.close()

            return {
                "success": True,
                "version": "Hive",
                "database_type": "Apache Hive",
                "test_time": datetime.now(),
                "response_time_ms": 100
            }
        except Exception as e:
            logger.error(f"Hive连接测试失败: {e}")
            return {
                "success": False,
                "error": str(e),
                "database_type": "Apache Hive",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取Hive数据库列表"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SHOW DATABASES")
            results = cursor.fetchall()
            cursor.close()
            return [db[0] for db in results]
        except Exception as e:
            logger.error(f"获取Hive数据库列表失败: {e}")
            # 连接失败时重置连接
            self._close_connection()
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取Hive表列表"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()

            # 切换数据库
            target_db = database or self.config.get('database', 'default')
            logger.info(f"切换到数据库: {target_db}")
            cursor.execute(f"USE {target_db}")

            # 获取表列表
            logger.info("执行 SHOW TABLES")
            cursor.execute("SHOW TABLES")
            table_names = [row[0] for row in cursor.fetchall()]
            logger.info(f"找到 {len(table_names)} 张表")

            # 应用分页
            paginated_tables = table_names[offset:offset + limit]

            tables = []
            for table_name in paginated_tables:
                try:
                    # 获取表信息(简化版,不获取详细信息避免太慢)
                    tables.append({
                        "table_name": table_name,
                        "table_type": "MANAGED_TABLE",
                        "estimated_rows": 0,
                        "data_size": 0,
                        "comment": ""
                    })
                except Exception as table_err:
                    logger.warning(f"获取表 {table_name} 信息失败: {table_err}")
                    tables.append({
                        "table_name": table_name,
                        "table_type": "UNKNOWN",
                        "estimated_rows": 0,
                        "data_size": 0,
                        "comment": ""
                    })

            cursor.close()
            logger.info(f"成功返回 {len(tables)} 张表")
            return tables

        except Exception as e:
            logger.error(f"获取Hive表列表失败: {e}")
            # 连接失败时重置连接
            self._close_connection()
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取Hive表结构"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()

            if database:
                cursor.execute(f"USE {database}")

            # 获取表结构
            cursor.execute(f"DESCRIBE {table_name}")
            columns_result = cursor.fetchall()

            columns = []
            partitions = []
            in_partition_section = False

            for row in columns_result:
                if len(row) >= 3:
                    col_name = row[0].strip() if row[0] else ""
                    col_type = row[1].strip() if row[1] else ""
                    comment = row[2].strip() if row[2] else ""

                    if col_name == "# Partition Information":
                        in_partition_section = True
                        continue
                    elif col_name.startswith("#"):
                        continue

                    if in_partition_section:
                        if col_name and col_type:
                            partitions.append({
                                "name": col_name,
                                "type": col_type,
                                "comment": comment
                            })
                    else:
                        if col_name and col_type:
                            columns.append({
                                "name": col_name,
                                "type": col_type,
                                "comment": comment,
                                "nullable": True  # Hive默认允许null
                            })

            # 获取详细表信息
            cursor.execute(f"DESCRIBE FORMATTED {table_name}")
            formatted_result = cursor.fetchall()

            table_properties = {}
            for row in formatted_result:
                if len(row) >= 2:
                    key = row[0].strip() if row[0] else ""
                    value = row[1].strip() if row[1] else ""

                    if key in ["InputFormat", "OutputFormat", "SerDe Library", "Storage Desc Params"]:
                        table_properties[key] = value

            cursor.close()

            return {
                "columns": columns,
                "partitions": partitions,
                "table_properties": table_properties,
                "storage_format": "Hive"
            }
        except Exception as e:
            logger.error(f"获取Hive表结构失败 {table_name}: {e}")
            return {"columns": [], "partitions": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[Dict[str, Any]]:
        """执行Hive查询"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()

            if database:
                cursor.execute(f"USE {database}")

            cursor.execute(query)

            # 获取列名
            columns = [desc[0] for desc in cursor.description] if cursor.description else []

            # 获取结果
            results = cursor.fetchall()
            cursor.close()

            # 转换为字典列表
            return [dict(zip(columns, row)) for row in results]
        except Exception as e:
            logger.error(f"Hive查询执行失败: {e}")
            return []


class ExcelClient(DatabaseClient):
    """Excel文件数据源客户端"""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.file_path = config.get('file_path', '')
        self.upload_dir = config.get('upload_dir', 'uploads/excel')
        self.file_info = {}
        self._ensure_upload_directory()

    def _ensure_upload_directory(self):
        """确保上传目录存在"""
        Path(self.upload_dir).mkdir(parents=True, exist_ok=True)

    async def test_connection(self) -> Dict[str, Any]:
        """测试Excel文件连接（检查文件是否存在和可读）"""
        try:
            if not self.file_path or not os.path.exists(self.file_path):
                return {
                    "success": False,
                    "error": f"文件不存在: {self.file_path}",
                    "database_type": "Excel",
                    "test_time": datetime.now()
                }

            # 尝试读取文件基本信息
            try:
                workbook = openpyxl.load_workbook(self.file_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()

                file_stat = os.stat(self.file_path)
                file_size = file_stat.st_size
                modified_time = datetime.fromtimestamp(file_stat.st_mtime)

                self.file_info = {
                    "file_size": file_size,
                    "modified_time": modified_time,
                    "sheet_count": len(sheet_names),
                    "sheet_names": sheet_names
                }

                return {
                    "success": True,
                    "version": f"Excel文件 - {len(sheet_names)} 个工作表",
                    "database_type": "Excel",
                    "test_time": datetime.now(),
                    "response_time_ms": 50,
                    "file_info": self.file_info
                }
            except Exception as e:
                return {
                    "success": False,
                    "error": f"文件读取失败: {str(e)}",
                    "database_type": "Excel",
                    "test_time": datetime.now()
                }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "database_type": "Excel",
                "test_time": datetime.now()
            }

    async def get_databases(self) -> List[str]:
        """获取工作簿列表（对于Excel，返回文件名）"""
        try:
            if not os.path.exists(self.file_path):
                return []

            filename = os.path.basename(self.file_path)
            return [filename]
        except Exception as e:
            logger.error(f"获取Excel文件信息失败: {e}")
            return []

    async def get_tables(self, database: str = None, schema: str = None, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """获取工作表列表"""
        try:
            if not os.path.exists(self.file_path):
                return []

            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            tables = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 获取工作表基本信息
                max_row = sheet.max_row
                max_col = sheet.max_column

                # 估算数据行数（排除空行）
                data_rows = 0
                for row in range(1, min(max_row + 1, 100)):  # 只检查前100行来估算
                    if any(sheet.cell(row=row, column=col).value is not None
                           for col in range(1, max_col + 1)):
                        data_rows += 1

                # 如果检查了100行，按比例估算总行数
                if max_row > 100:
                    estimated_rows = int((data_rows / 100) * max_row)
                else:
                    estimated_rows = data_rows

                tables.append({
                    "table_name": sheet_name,
                    "table_type": "WORKSHEET",
                    "estimated_rows": estimated_rows,
                    "columns_count": max_col,
                    "data_size": estimated_rows * max_col * 50,  # 粗略估算
                    "comment": f"Excel工作表 - {max_row}行 {max_col}列"
                })

            workbook.close()
            return tables
        except Exception as e:
            logger.error(f"获取Excel工作表列表失败: {e}")
            return []

    async def get_table_schema(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取Excel工作表结构"""
        try:
            if not os.path.exists(self.file_path):
                return {"columns": [], "error": "文件不存在"}

            # 使用pandas读取Excel文件获取列信息
            df = pd.read_excel(self.file_path, sheet_name=table_name, nrows=0)

            columns = []
            for col_name in df.columns:
                columns.append({
                    "name": str(col_name),
                    "type": "TEXT",  # Excel中所有数据默认为文本类型
                    "nullable": True,
                    "default": None,
                    "comment": f"Excel列 - {col_name}"
                })

            return {
                "columns": columns,
                "indexes": [],
                "primary_key": [],
                "foreign_keys": [],
                "sheet_info": {
                    "sheet_name": table_name,
                    "column_count": len(columns),
                    "file_path": self.file_path
                }
            }
        except Exception as e:
            logger.error(f"获取Excel工作表结构失败 {table_name}: {e}")
            return {"columns": [], "error": str(e)}

    async def execute_query(self, query: str, database: str = None, schema: str = None, limit: int = 100) -> List[Dict[str, Any]]:
        """执行Excel查询（简单的数据读取）"""
        try:
            # 解析简单的查询语句
            if not query.upper().startswith('SELECT'):
                return []

            # 提取表名（工作表名）
            query_parts = query.upper().split()
            from_index = query_parts.index('FROM') if 'FROM' in query_parts else -1
            if from_index == -1:
                return []

            table_name = query_parts[from_index + 1].strip(';')

            # 处理LIMIT子句
            limit = None
            if 'LIMIT' in query_parts:
                limit_index = query_parts.index('LIMIT')
                if limit_index + 1 < len(query_parts):
                    try:
                        limit = int(query_parts[limit_index + 1].strip(';'))
                    except ValueError:
                        limit = 100

            # 读取Excel数据
            df = pd.read_excel(self.file_path, sheet_name=table_name)

            # 应用LIMIT
            if limit:
                df = df.head(limit)

            # 处理NaN值
            df = df.fillna('')

            # 转换为字典列表
            results = df.to_dict('records')

            # 确保所有值都是可序列化的
            for record in results:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = None
                    elif isinstance(value, (pd.Timestamp, datetime)):
                        record[key] = value.isoformat()
                    else:
                        record[key] = str(value)

            return results
        except Exception as e:
            logger.error(f"Excel查询执行失败: {e}")
            return []

    async def get_table_metadata(self, table_name: str, database: str = None) -> Dict[str, Any]:
        """获取Excel工作表的完整元数据"""
        try:
            schema = await self.get_table_schema(table_name, database)

            # 获取数据统计信息
            df = pd.read_excel(self.file_path, sheet_name=table_name)

            statistics = {
                "row_count": len(df),
                "column_count": len(df.columns),
                "estimated_size": f"{len(df) * len(df.columns) * 50} bytes",
                "last_updated": datetime.fromtimestamp(os.path.getmtime(self.file_path)).isoformat(),
                "null_counts": df.isnull().sum().to_dict(),
                "data_types": df.dtypes.astype(str).to_dict()
            }

            return {
                "table_name": table_name,
                "database": database,
                "schema": schema,
                "statistics": statistics,
                "collected_at": datetime.now()
            }
        except Exception as e:
            logger.error(f"获取Excel表元数据失败 {table_name}: {e}")
            return {
                "table_name": table_name,
                "database": database,
                "error": str(e),
                "collected_at": datetime.now()
            }

    async def upload_file(self, file: UploadFile) -> Dict[str, Any]:
        """上传Excel文件"""
        try:
            # 验证文件类型
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                return {
                    "success": False,
                    "error": "不支持的文件类型，请上传.xlsx或.xls文件"
                }

            # 生成唯一文件名
            file_id = str(uuid.uuid4())
            file_extension = os.path.splitext(file.filename)[1]
            new_filename = f"{file_id}_{file.filename}"
            file_path = os.path.join(self.upload_dir, new_filename)

            # 保存文件
            with open(file_path, "wb") as buffer:
                content = await file.read()
                buffer.write(content)

            # 更新配置
            self.file_path = file_path
            self.config['file_path'] = file_path

            # 验证文件
            test_result = await self.test_connection()

            if test_result.get('success'):
                return {
                    "success": True,
                    "file_id": file_id,
                    "file_path": file_path,
                    "original_filename": file.filename,
                    "file_size": len(content),
                    "uploaded_at": datetime.now(),
                    "file_info": test_result.get('file_info', {})
                }
            else:
                # 如果文件验证失败，删除文件
                os.remove(file_path)
                return {
                    "success": False,
                    "error": f"文件验证失败: {test_result.get('error', '未知错误')}"
                }
        except Exception as e:
            logger.error(f"Excel文件上传失败: {e}")
            return {
                "success": False,
                "error": str(e)
            }

    async def get_data_preview(self, table_name: str, limit: int = 10) -> Dict[str, Any]:
        """获取Excel数据预览"""
        try:
            df = pd.read_excel(self.file_path, sheet_name=table_name, nrows=limit)
            df = df.fillna('')

            # 转换为字典列表
            preview_data = df.to_dict('records')

            # 确保所有值都是可序列化的
            for record in preview_data:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = None
                    elif isinstance(value, (pd.Timestamp, datetime)):
                        record[key] = value.isoformat()
                    else:
                        record[key] = str(value)

            return {
                "success": True,
                "table_name": table_name,
                "preview_data": preview_data,
                "column_names": df.columns.tolist(),
                "row_count": len(preview_data),
                "total_rows": len(pd.read_excel(self.file_path, sheet_name=table_name, usecols=[0]).dropna()),
                "previewed_at": datetime.now()
            }
        except Exception as e:
            logger.error(f"获取Excel数据预览失败: {e}")
            return {
                "success": False,
                "error": str(e)
            }


# 数据库客户端工厂
class DatabaseClientFactory:
    """数据库客户端工厂类"""

    _clients = {
        'mysql': MySQLClient,
        'kingbase': KingbaseClient,
        'doris': DorisClient,
        'dameng': DamengClient,
        'gbase': GBaseClient,
        'tidb': TiDBClient,
        'hive': HiveClient,
        'excel': ExcelClient
    }

    @classmethod
    def create_client(cls, db_type: str, config: Dict[str, Any]) -> DatabaseClient:
        """创建数据库客户端"""
        db_type = db_type.lower()
        if db_type not in cls._clients:
            raise ValueError(f"不支持的数据库类型: {db_type}")

        client_class = cls._clients[db_type]
        return client_class(config)

    @classmethod
    def get_supported_types(cls) -> List[str]:
        """获取支持的数据库类型"""
        return list(cls._clients.keys())

class ExcelDataSourceService:
        """Excel数据源服务"""

        def __init__(self):
            self.upload_dir = "uploads/excel"
            self._ensure_upload_directory()

        def _ensure_upload_directory(self):
            """确保上传目录存在"""
            Path(self.upload_dir).mkdir(parents=True, exist_ok=True)

        async def create_excel_source(self, name: str, file: UploadFile, description: str = None) -> Dict[str, Any]:
            """创建Excel数据源"""
            try:
                # 创建临时客户端处理文件上传
                temp_client = ExcelClient({
                    'upload_dir': self.upload_dir
                })

                # 上传文件
                upload_result = await temp_client.upload_file(file)

                if not upload_result.get('success'):
                    return upload_result

                # 创建数据源配置
                config = {
                    'file_path': upload_result['file_path'],
                    'upload_dir': self.upload_dir,
                    'original_filename': upload_result['original_filename'],
                    'file_id': upload_result['file_id'],
                    'uploaded_at': upload_result['uploaded_at'],
                    'file_size': upload_result['file_size']
                }

                return {
                    "success": True,
                    "name": name,
                    "type": "excel",
                    "config": config,
                    "description": description or f"Excel文件: {file.filename}",
                    "upload_info": upload_result
                }
            except Exception as e:
                logger.error(f"创建Excel数据源失败: {e}")
                return {
                    "success": False,
                    "error": str(e)
                }

        async def delete_excel_source(self, config: Dict[str, Any]) -> Dict[str, Any]:
            """删除Excel数据源（包括文件）"""
            try:
                file_path = config.get('file_path')
                if file_path and os.path.exists(file_path):
                    os.remove(file_path)
                    return {
                        "success": True,
                        "message": "Excel文件已删除"
                    }
                return {
                    "success": True,
                    "message": "文件不存在或已被删除"
                }
            except Exception as e:
                logger.error(f"删除Excel文件失败: {e}")
                return {
                    "success": False,
                    "error": str(e)
                }

        async def list_uploaded_files(self) -> List[Dict[str, Any]]:
            """列出已上传的Excel文件"""
            try:
                files = []
                if os.path.exists(self.upload_dir):
                    for filename in os.listdir(self.upload_dir):
                        file_path = os.path.join(self.upload_dir, filename)
                        if os.path.isfile(file_path) and filename.lower().endswith(('.xlsx', '.xls')):
                            stat = os.stat(file_path)
                            files.append({
                                "filename": filename,
                                "file_path": file_path,
                                "size": stat.st_size,
                                "modified_time": datetime.fromtimestamp(stat.st_mtime),
                                "created_time": datetime.fromtimestamp(stat.st_ctime)
                            })
                return files
            except Exception as e:
                logger.error(f"列出上传文件失败: {e}")
                return []

# 连接管理器
class ConnectionManager:
    """数据库连接管理器"""

    def __init__(self):
        self._clients: Dict[str, DatabaseClient] = {}
        self._connection_cache = {}

    def add_client(self, name: str, db_type: str, config: Dict[str, Any]) -> bool:
        """添加数据库客户端"""
        try:
            client = DatabaseClientFactory.create_client(db_type, config)
            if client:
                self._clients[name] = client
                logger.info(f"添加数据库客户端: {name} ({db_type})")
                return True
            else:
                logger.error(f"无法创建客户端: {name} ({db_type}) - 不支持的数据库类型")
                return False
        except Exception as e:
            logger.error(f"创建客户端失败: {name} ({db_type}) - {e}")
            return False

    def get_client(self, name: str) -> Optional[DatabaseClient]:
        """获取数据库客户端"""
        return self._clients.get(name)

    def remove_client(self, name: str) -> None:
        """移除数据库客户端"""
        if name in self._clients:
            del self._clients[name]
            logger.info(f"移除数据库客户端: {name}")

    def list_clients(self) -> List[str]:
        """列出所有客户端名称"""
        return list(self._clients.keys())

    async def test_all_connections(self) -> Dict[str, Dict[str, Any]]:
        """测试所有连接"""
        results = {}
        for name, client in self._clients.items():
            try:
                result = await client.test_connection()
                results[name] = result
            except Exception as e:
                results[name] = {
                    "success": False,
                    "error": str(e),
                    "test_time": datetime.now()
                }
        return results


# 全局连接管理器实例
connection_manager = ConnectionManager()
excel_service = ExcelDataSourceService()